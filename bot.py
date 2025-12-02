import os
import telebot
from telebot import types
from dotenv import load_dotenv
import json
from datetime import datetime, timedelta
import threading
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
load_dotenv()

BOT_TOKEN = os.getenv('BOT_TOKEN')
SUBMIT_LINK = os.getenv('SUBMIT_LINK', 'https://clubgg.app.link/your-link')
PAYMENT_LINK = os.getenv('PAYMENT_LINK', 'https://t.me/sapayobot')
COMMUNITY_LINK = os.getenv('COMMUNITY_LINK', 'https://t.me/+gl35BSf3cBY0YjUy')
MANAGER_LINK = os.getenv('MANAGER_LINK', 'https://t.me/karinakroks')
MENU_IMAGE = os.getenv('MENU_IMAGE', 'меню.png')
ADMIN_ID = 6854574142  # Главный админ ID
ADMINS_FILE = 'admins.json'  # Файл для хранения всех админов
LANGUAGE_FILE = 'languages.json'
SUPPORT_SETTINGS_FILE = 'support_settings.json'  # Настройки поддержки админов
SUPPORT_TICKETS_FILE = 'support_tickets.json'  # Активные тикеты поддержки
APPLICATIONS_FILE = 'applications.json'  # Заявки на присоединение
RECEIPTS_FILE = 'receipts.json'  # Чеки на зачисление денег
PROMO_CODES_FILE = 'promo_codes.json'  # Промокоды для пополнения баланса
SETTINGS_FILE = 'settings.json'  # Файл с переменными конфигурации
SERVICE_NAME = 'PokerKingClubBot'  # Имя сервиса для платежей
REFERRAL_PERCENT = 0.10  # 10% процентов рефереру
SUPPORT_COOLDOWN_SECONDS = 30  # Минимум 30 секунд между запросами
MAX_ACTIVE_TICKETS_PER_USER = 1  # Максимум 1 активный тикет на пользователя

if not BOT_TOKEN:
    raise RuntimeError('Please set BOT_TOKEN in environment or .env file')

bot = telebot.TeleBot(BOT_TOKEN)
user_states = {}
user_messages = {}  # Словарь для отслеживания последних сообщений пользователей {user_id: [message_ids]}
BALANCE_FILE = 'balances.json'
INIT_STATE_FILE = 'init_states.json'  # Для отслеживания инициализации пользователей
ACTIONS_FILE = 'user_actions.json'  # Для отслеживания действий пользователей
AUTO_DELETE_TIMEOUT = 60  # Время автоудаления сообщений в секундах (60 сек = 1 минута)

# Защита от DDoS - отслеживание попыток создания тикетов
support_requests_cooldown = {}  # {user_id: timestamp}
SUPPORT_COOLDOWN_SECONDS = 30  # Минимум 30 секунд между запросами
MAX_ACTIVE_TICKETS_PER_USER = 1  # Максимум 1 активный тикет на пользователя

# Функции для управления админами
def load_admins():
    """Загружает админов с их типами"""
    if os.path.exists(ADMINS_FILE):
        try:
            with open(ADMINS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Если это старый формат - миграция
                if 'admins' in data and isinstance(data['admins'], list):
                    # Конвертируем старый формат
                    admins_dict = {}
                    for admin_id in data['admins']:
                        admins_dict[str(admin_id)] = 'full' if admin_id == ADMIN_ID else 'full'
                    return admins_dict
                return data.get('admins', {str(ADMIN_ID): 'full'})
        except:
            return {str(ADMIN_ID): 'full'}
    return {str(ADMIN_ID): 'full'}


def save_admins(admins_dict):
    """Сохраняет админов с их типами"""
    with open(ADMINS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'admins': admins_dict}, f, ensure_ascii=False, indent=2)


def is_admin(user_id):
    """Проверяет является ли пользователь админом"""
    admins = load_admins()
    return str(user_id) in admins


def get_admin_type(user_id):
    """Получает тип админа (full или support)"""
    admins = load_admins()
    admin_type = admins.get(str(user_id), None)
    return admin_type


def is_full_admin(user_id):
    """Проверяет является ли пользователь полным админом"""
    return get_admin_type(user_id) == 'full'


def is_support_admin(user_id):
    """Проверяет является ли пользователь админом поддержки"""
    return get_admin_type(user_id) == 'support'


def add_admin(user_id, admin_type='full'):
    """Добавляет админа"""
    if user_id == ADMIN_ID:
        return False  # Главный админ уже существует
    
    admins = load_admins()
    
    if str(user_id) in admins:
        return False  # Админ уже существует
    
    admins[str(user_id)] = admin_type
    save_admins(admins)
    return True


def remove_admin(user_id):
    """Удаляет админа"""
    if user_id == ADMIN_ID:
        return False  # Не можем удалить главного админа
    
    admins = load_admins()
    
    if str(user_id) not in admins:
        return False
    
    del admins[str(user_id)]
    save_admins(admins)
    return True


# Функции для системы поддержки
def load_support_settings():
    """Загружает настройки поддержки админов"""
    if os.path.exists(SUPPORT_SETTINGS_FILE):
        try:
            with open(SUPPORT_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_support_settings(settings):
    """Сохраняет настройки поддержки"""
    with open(SUPPORT_SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def load_support_tickets():
    """Загружает активные тикеты поддержки"""
    if os.path.exists(SUPPORT_TICKETS_FILE):
        try:
            with open(SUPPORT_TICKETS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_support_tickets(tickets):
    """Сохраняет тикеты поддержки"""
    with open(SUPPORT_TICKETS_FILE, 'w', encoding='utf-8') as f:
        json.dump(tickets, f, ensure_ascii=False, indent=2)


def get_admin_name(admin_id):
    """Получает имя админа"""
    try:
        member = bot.get_chat_member(-1001234567890, admin_id)  # Примерный ID группы
        return member.user.first_name or "Администратор"
    except:
        return "Администратор"


def is_support_enabled(admin_id):
    """Проверяет, включена ли поддержка для админа"""
    settings = load_support_settings()
    return settings.get(str(admin_id), False)


# Функции для системы заявок
def load_applications():
    """Загружает заявки"""
    if os.path.exists(APPLICATIONS_FILE):
        try:
            with open(APPLICATIONS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_applications(applications):
    """Сохраняет заявки"""
    with open(APPLICATIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(applications, f, ensure_ascii=False, indent=2)


def load_applications_settings():
    """Загружает настройки заявок (кто принимает)"""
    settings = load_support_settings()
    # Используем поле support_enabled для заявок тоже
    return settings


def is_applications_enabled(admin_id):
    """Проверяет, включен ли приём заявок у админа"""
    settings = load_applications_settings()
    return settings.get(str(admin_id), False)


# Функции для работы с чеками
def load_receipts():
    """Загружает чеки на зачисление"""
    if os.path.exists(RECEIPTS_FILE):
        try:
            with open(RECEIPTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_receipts(receipts):
    """Сохраняет чеки"""
    with open(RECEIPTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(receipts, f, ensure_ascii=False, indent=2)


def generate_game_id():
    """Генерирует уникальный game_id в формате XXXX-XXXX"""
    import random
    part1 = ''.join([str(random.randint(0, 9)) for _ in range(4)])
    part2 = ''.join([str(random.randint(0, 9)) for _ in range(4)])
    return f"{part1}-{part2}"


# Функции для работы с промокодами
def load_promo_codes():
    """Загружает промокоды"""
    if os.path.exists(PROMO_CODES_FILE):
        try:
            with open(PROMO_CODES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_promo_codes(codes):
    """Сохраняет промокоды"""
    with open(PROMO_CODES_FILE, 'w', encoding='utf-8') as f:
        json.dump(codes, f, ensure_ascii=False, indent=2)


def generate_promo_code(amount):
    """Генерирует уникальный промокод"""
    import random
    import string
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    return code


def create_promo_codes(amount, quantity):
    """Создает n промокодов на заданную сумму"""
    codes = load_promo_codes()
    created_codes = []
    
    for _ in range(quantity):
        code = generate_promo_code(amount)
        # Избегаем дубликатов
        while code in codes:
            code = generate_promo_code(amount)
        
        codes[code] = {
            'amount': amount,
            'used_by': None,
            'used_at': None,
            'created_at': datetime.now().isoformat(),
            'status': 'active'
        }
        created_codes.append(code)
    
    save_promo_codes(codes)
    return created_codes


# Функции для управления настройками конфигурации
def load_settings():
    """Загружает настройки конфигурации"""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return get_default_settings()
    return get_default_settings()


def save_settings(settings):
    """Сохраняет настройки конфигурации"""
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def get_default_settings():
    """Возвращает настройки по умолчанию"""
    return {
        'ADMIN_ID': ADMIN_ID,
        'RECEIPT_AGENT_ID': ADMIN_ID,
        'SERVICE_NAME': SERVICE_NAME,
        'REFERRAL_PERCENT': REFERRAL_PERCENT,
        'REFERRAL_THRESHOLD': 50000,
        'SUPPORT_COOLDOWN_SECONDS': SUPPORT_COOLDOWN_SECONDS,
        'MAX_ACTIVE_TICKETS_PER_USER': MAX_ACTIVE_TICKETS_PER_USER,
        'SUBMIT_LINK': SUBMIT_LINK,
        'PAYMENT_LINK': PAYMENT_LINK,
        'COMMUNITY_LINK': COMMUNITY_LINK,
        'MANAGER_LINK': MANAGER_LINK,
        'MENU_IMAGE': MENU_IMAGE,
        'AUTO_DELETE_TIMEOUT': AUTO_DELETE_TIMEOUT,
    }


def get_setting(key, default=None):
    """Получает значение настройки"""
    settings = load_settings()
    return settings.get(key, default)


def set_setting(key, value):
    """Устанавливает значение настройки"""
    settings = load_settings()
    settings[key] = value
    save_settings(settings)

# Языковые данные
TRANSLATIONS = {
    'ru': {
        'welcome': '👋 <b>Добро пожаловать в Kingxxx Club!</b>',
        'apply': '✅ Подать заявку',
        'promos': '🎁 Акции/Бонусы',
        'payment': '💳 Оплата',
        'manager': '👔 Менеджер',
        'community': '💬 Сообщество',
        'referral_prog': '👥 Реферальная программа',
        'profile': '👤 Личный кабинет',
        'language': '🌐 Язык',
        'back': '⬅️ Назад в меню',
        'profile_title': '<b>👤 ЛИЧНЫЙ КАБИНЕТ</b> 👤',
        'balance': 'Баланс',
        'referrals': 'Рефералов приглашено',
        'earned': 'Заработано',
        'member_since': 'Участник с',
        'profile_back': '⬅️ Назад',
        'ref_link': '📤 Скопировать реферальную ссылку',
        'ref_stats': '📊 Статистика',
        'promos_title': '🎁 АКЦИИ МЕСЯЦА 🎁',
        'promos_badbeat': '🔥 БАД БИТ\n   Проиграл сильную руку? Компенсация автоматически!',
        'promos_freeroll': '🎰 ФРИСТАРТ\n   Бесплатные турниры + призы для новичков!',
        'promos_referral': '💵 ПАССИВНЫЙ ДОХОД\n   Пригласи друга → Получай постоянно!',
        'ref_stats_title': '📊 СТАТИСТИКА РЕФЕРАЛОВ',
        'ref_stats_invited': 'Всего приглашено',
        'ref_stats_earned': 'Примерный доход',
        'ref_stats_message': 'Приглашай больше друзей и\nполучай больше награды! 🚀',
        'referral_title': '👥 РЕФЕРАЛЬНАЯ ПРОГРАММА',
        'referral_how': 'КАК ЭТО РАБОТАЕТ?',
        'referral_step1': '1️⃣ Получи персональную ссылку',
        'referral_step2': '2️⃣ Пригласи друзей по этой ссылке',
        'referral_step3': '3️⃣ Получай комиссию за каждого',
        'referral_reward': 'РАЗМЕР НАГРАДЫ',
        'referral_reward_desc': '➜ 10% от первого бай-ина друга',
        'referral_benefits': 'ПРЕИМУЩЕСТВА',
        'referral_benefit1': '✅ Неограниченный доход',
        'referral_benefit2': '✅ Быстрое начисление',
        'referral_benefit3': '✅ Никаких комиссий',
        'referral_click': 'Нажми кнопку ниже для ссылки!',
    },
    'en': {
        'welcome': '👋 <b>Welcome to Kinxxx Club!</b>',
        'apply': '✅ Submit Application',
        'promos': '🎁 Promos/Bonuses',
        'payment': '💳 Payment',
        'manager': '👔 Manager',
        'community': '💬 Community',
        'referral_prog': '👥 Referral Program',
        'profile': '👤 Personal Cabinet',
        'language': '🌐 Language',
        'back': '⬅️ Back to Menu',
        'profile_title': '<b>👤 PERSONAL CABINET</b> 👤',
        'balance': 'Balance',
        'referrals': 'Referrals Invited',
        'earned': 'Earned',
        'member_since': 'Member Since',
        'profile_back': '⬅️ Back',
        'ref_link': '📤 Copy Link',
        'ref_stats': '📊 Statistics',
        'promos_title': '🎁 PROMOTIONS OF THE MONTH 🎁',
        'promos_badbeat': '🔥 BAD BEAT\n   Unlucky hand? Get compensation instantly!',
        'promos_freeroll': '🎰 FREE START\n   Free tournaments + instant prizes!',
        'promos_referral': '💵 PASSIVE INCOME\n   Invite a friend → Earn forever!',
        'ref_stats_title': '📊 REFERRAL STATISTICS',
        'ref_stats_invited': 'Total invited',
        'ref_stats_earned': 'Estimated income',
        'ref_stats_message': 'Invite more friends and\nget more rewards! 🚀',
        'referral_title': '👥 REFERRAL PROGRAM',
        'referral_how': 'HOW DOES IT WORK?',
        'referral_step1': '1️⃣ Get your personal link',
        'referral_step2': '2️⃣ Invite friends via this link',
        'referral_step3': '3️⃣ Get commission for each one',
        'referral_reward': 'REWARD SIZE',
        'referral_reward_desc': '➜ 10% from friend\'s first buy-in',
        'referral_benefits': 'BENEFITS',
        'referral_benefit1': '✅ Unlimited income',
        'referral_benefit2': '✅ Quick payouts',
        'referral_benefit3': '✅ No fees',
        'referral_click': 'Click the button below for the link!',
    }
}


def load_user_language(user_id):
    if os.path.exists(LANGUAGE_FILE):
        try:
            with open(LANGUAGE_FILE, 'r', encoding='utf-8') as f:
                langs = json.load(f)
                return langs.get(str(user_id), 'ru')
        except:
            return 'ru'
    return 'ru'


def save_user_language(user_id, lang):
    try:
        if os.path.exists(LANGUAGE_FILE):
            with open(LANGUAGE_FILE, 'r', encoding='utf-8') as f:
                langs = json.load(f)
        else:
            langs = {}
    except:
        langs = {}
    
    langs[str(user_id)] = lang
    with open(LANGUAGE_FILE, 'w', encoding='utf-8') as f:
        json.dump(langs, f, ensure_ascii=False, indent=2)


def get_text(user_id, key):
    lang = load_user_language(user_id)
    return TRANSLATIONS.get(lang, TRANSLATIONS['ru']).get(key, key)

# Файлы для хранения данных
REFERRAL_FILE = 'referrals.json'
USERS_FILE = 'users.json'


def load_balances():
    if os.path.exists(BALANCE_FILE):
        try:
            with open(BALANCE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_balances(data):
    with open(BALANCE_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_balance(user_id):
    balances = load_balances()
    return balances.get(str(user_id), 0)


def set_balance(user_id, amount):
    balances = load_balances()
    balances[str(user_id)] = amount
    save_balances(balances)


def add_balance(user_id, amount):
    current = get_balance(user_id)
    set_balance(user_id, current + amount)


def delete_message_after_timeout(chat_id, message_id, timeout=AUTO_DELETE_TIMEOUT):
    """Удаляет сообщение через указанное время (в секундах)"""
    def delete():
        try:
            time.sleep(timeout)
            bot.delete_message(chat_id, message_id)
        except:
            pass  # Сообщение уже удалено или другая ошибка
    
    thread = threading.Thread(target=delete, daemon=True)
    thread.start()


def delete_old_messages(chat_id):
    """Удаляет все старые сообщения пользователя, оставляя только новое"""
    if chat_id in user_messages:
        old_messages = user_messages[chat_id]
        for msg_id in old_messages:
            try:
                bot.delete_message(chat_id, msg_id)
            except:
                pass  # Сообщение уже удалено или другая ошибка
        user_messages[chat_id] = []


def track_message(chat_id, message_id):
    """Отслеживает новое сообщение от бота"""
    delete_old_messages(chat_id)  # Удаляем старые сообщения
    if chat_id not in user_messages:
        user_messages[chat_id] = []
    user_messages[chat_id].append(message_id)


def load_init_states():
    """Загружает состояние инициализации пользователей"""
    if os.path.exists(INIT_STATE_FILE):
        try:
            with open(INIT_STATE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_init_states(data):
    """Сохраняет состояние инициализации пользователей"""
    with open(INIT_STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_init_state(user_id):
    """Получает состояние инициализации пользователя"""
    states = load_init_states()
    return states.get(str(user_id), 'not_started')


def set_init_state(user_id, state):
    """Устанавливает состояние инициализации пользователя"""
    states = load_init_states()
    states[str(user_id)] = state
    save_init_states(states)


# ==================== СИСТЕМА ОТСЛЕЖИВАНИЯ ДЕЙСТВИЙ ====================

def load_user_actions():
    """Загружает действия пользователей"""
    if os.path.exists(ACTIONS_FILE):
        try:
            with open(ACTIONS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_user_actions(data):
    """Сохраняет действия пользователей"""
    with open(ACTIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def track_action(user_id, action_type):
    """Отслеживает действие пользователя"""
    actions = load_user_actions()
    user_id_str = str(user_id)
    
    if user_id_str not in actions:
        actions[user_id_str] = {}
    
    # Записываем действие
    if action_type not in actions[user_id_str]:
        actions[user_id_str][action_type] = 0
    
    actions[user_id_str][action_type] += 1
    save_user_actions(actions)


def get_user_exploration_stats(user_id):
    """Получает статистику изучения бота пользователем"""
    actions = load_user_actions()
    user_id_str = str(user_id)
    user_actions = actions.get(user_id_str, {})
    
    # Определяем все возможные кнопки
    all_buttons = [
        'promos', 'bonus', 'referral', 'copy_ref', 'ref_stats',
        'profile', 'language', 'lang_ru', 'lang_en',
        'payment', 'community', 'manager', 'menu'
    ]
    
    # Считаем количество нажатых кнопок
    clicked_buttons = [btn for btn in all_buttons if btn in user_actions]
    clicked_count = len(clicked_buttons)
    total_buttons = len(all_buttons)
    
    # Считаем процент
    exploration_percent = int((clicked_count / total_buttons) * 100) if total_buttons > 0 else 0
    
    return {
        'percent': exploration_percent,
        'clicked': clicked_count,
        'total': total_buttons,
        'actions': user_actions
    }


def load_referrals():
    if os.path.exists(REFERRAL_FILE):
        try:
            with open(REFERRAL_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_referrals(data):
    with open(REFERRAL_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_users(data):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_referral_link(user_id):
    return f"https://t.me/{bot.get_me().username}?start=ref_{user_id}"


def main_menu_kbd(user_id):
    lang = load_user_language(user_id)
    kb = types.InlineKeyboardMarkup()
    
    # Первая строка: Профиль (2х1)
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'profile'), callback_data='profile'))
    
    # Вторая строка: Рефералы (2х1)
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'referral_prog'), callback_data='referral'))
    
    # Третья строка: Оплата (2х1)
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'payment'), callback_data='payment_info'))
    
    # Четвертая строка: Сообщество (2х1)
    kb.add(types.InlineKeyboardButton('👥 ' + ('Community' if lang == 'en' else 'Сообщество'), callback_data='community_menu'))
    
    # Пятая строка: Язык (2х1)
    kb.add(types.InlineKeyboardButton('🌐 ' + ('Language' if lang == 'en' else 'Язык'), callback_data='language'))
    
    return kb


def community_menu_kbd(user_id):
    """Меню сообщества с тремя кнопками"""
    lang = load_user_language(user_id)
    kb = types.InlineKeyboardMarkup()
    
    if lang == 'en':
        kb.add(types.InlineKeyboardButton('✅ Submit Application', callback_data='apply_conditions'))
        kb.add(types.InlineKeyboardButton('🎁 Promotions', callback_data='promos'))
        kb.add(types.InlineKeyboardButton('👔 Manager', url=MANAGER_LINK))
    else:
        kb.add(types.InlineKeyboardButton('✅ Подать заявку', callback_data='apply_conditions'))
        kb.add(types.InlineKeyboardButton('🎁 Акции', callback_data='promos'))
        kb.add(types.InlineKeyboardButton('👔 Связь с менеджером', url=MANAGER_LINK))
    
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'back'), callback_data='menu'))
    return kb


def apply_conditions_kbd(user_id):
    """Меню подать заявку с тремя кнопками"""
    lang = load_user_language(user_id)
    kb = types.InlineKeyboardMarkup()
    
    if lang == 'en':
        kb.add(types.InlineKeyboardButton('💬 Join Chat', url=COMMUNITY_LINK))
        kb.add(types.InlineKeyboardButton('📋 Submit Application', callback_data='submit_application'))
        kb.add(types.InlineKeyboardButton('⚖️ Terms & Conditions', callback_data='show_conditions'))
    else:
        kb.add(types.InlineKeyboardButton('💬 Присоединиться к чату', url=COMMUNITY_LINK))
        kb.add(types.InlineKeyboardButton('📋 Подать заявку в клуб', callback_data='submit_application'))
        kb.add(types.InlineKeyboardButton('⚖️ Общие условия', callback_data='show_conditions'))
    
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'back'), callback_data='community_menu_back'))
    return kb


def promos_kbd(user_id):
    lang = load_user_language(user_id)
    kb = types.InlineKeyboardMarkup()
    if lang == 'en':
        kb.add(types.InlineKeyboardButton('📋 Learn Conditions', callback_data='bonus'))
        kb.add(types.InlineKeyboardButton('✅ Submit Application to Club', callback_data='apply'))
    else:
        kb.add(types.InlineKeyboardButton('📋 Узнать условия', callback_data='bonus'))
        kb.add(types.InlineKeyboardButton('✅ Подать заявку в клуб', callback_data='apply'))
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'back'), callback_data='menu'))
    return kb


def referral_kbd(user_id):
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'ref_link'), callback_data='copy_ref'))
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'ref_stats'), callback_data='ref_stats'))
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'back'), callback_data='menu'))
    return kb


def back_kbd(user_id):
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'back'), callback_data='menu'))
    return kb


def bonus_kbd(user_id):
    lang = load_user_language(user_id)
    kb = types.InlineKeyboardMarkup()
    if lang == 'en':
        kb.add(types.InlineKeyboardButton('✅ Submit Application to Club', callback_data='apply'))
    else:
        kb.add(types.InlineKeyboardButton('✅ Подать заявку в клуб', callback_data='apply'))
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'back'), callback_data='promos'))
    return kb


def language_kbd():
    kb = types.InlineKeyboardMarkup()
    kb.add(
        types.InlineKeyboardButton('🇷🇺 Русский', callback_data='lang_ru'),
        types.InlineKeyboardButton('🇬🇧 English', callback_data='lang_en'),
    )
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='menu'))
    return kb


def profile_kbd(user_id):
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'ref_link'), callback_data='copy_ref'))
    kb.add(types.InlineKeyboardButton(get_text(user_id, 'profile_back'), callback_data='menu'))
    return kb


def admin_menu_kbd(admin_id=None):
    kb = types.InlineKeyboardMarkup()
    
    # Если это админ поддержки - показываем только его меню
    if admin_id and is_support_admin(admin_id):
        kb.add(types.InlineKeyboardButton('📞 Настройки поддержки', callback_data='admin_support_settings'))
        kb.add(types.InlineKeyboardButton('⬅️ Выход', callback_data='admin_exit'))
        return kb
    
    # Полное меню для полных админов
    kb.add(types.InlineKeyboardButton('💰 Управление балансом', callback_data='admin_balance'))
    kb.add(types.InlineKeyboardButton('👥 Управление пользователями', callback_data='admin_users'))
    kb.add(types.InlineKeyboardButton('📋 Все пользователи', callback_data='admin_all_users'))
    kb.add(types.InlineKeyboardButton('📊 Статистика', callback_data='admin_stats'))
    kb.add(types.InlineKeyboardButton('🔍 Изучение пользователей', callback_data='admin_exploration'))
    kb.add(types.InlineKeyboardButton('📢 Рассылка всем', callback_data='admin_broadcast'))
    kb.add(types.InlineKeyboardButton('🔑 Управление админами', callback_data='admin_manage_admins'))
    kb.add(types.InlineKeyboardButton('📞 Настройки поддержки', callback_data='admin_support_settings'))
    kb.add(types.InlineKeyboardButton('⚙️ Настройки переменных', callback_data='admin_settings_variables'))
    kb.add(types.InlineKeyboardButton('📥 Скачать информацию', callback_data='admin_download_menu'))
    kb.add(types.InlineKeyboardButton('⬅️ Выход', callback_data='admin_exit'))
    return kb


def admin_balance_kbd():
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('💳 Пополнить баланс', callback_data='admin_recharge_balance'))
    kb.add(types.InlineKeyboardButton('➕ Добавить баланс', callback_data='admin_add_balance'))
    kb.add(types.InlineKeyboardButton('➖ Убрать баланс', callback_data='admin_remove_balance'))
    kb.add(types.InlineKeyboardButton('🔧 Установить баланс', callback_data='admin_set_balance'))
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    return kb


def send_menu(chat_id):
    text = get_text(chat_id, 'welcome')
    kb = main_menu_kbd(chat_id)
    try:
        if MENU_IMAGE and (MENU_IMAGE.startswith('http') or os.path.exists(MENU_IMAGE)):
            if MENU_IMAGE.startswith('http'):
                msg = bot.send_photo(chat_id, photo=MENU_IMAGE, caption=text, reply_markup=kb, parse_mode='HTML')
                track_message(chat_id, msg.message_id)
            else:
                with open(MENU_IMAGE, 'rb') as photo:
                    msg = bot.send_photo(chat_id, photo=photo, caption=text, reply_markup=kb, parse_mode='HTML')
                    track_message(chat_id, msg.message_id)
            return
    except Exception:
        pass
    msg = bot.send_message(chat_id, text, reply_markup=kb, parse_mode='HTML')
    track_message(chat_id, msg.message_id)


def send_profile(chat_id):
    users = load_users()
    referrals_data = load_referrals()
    user_info = users.get(str(chat_id), {})
    user_referrals = referrals_data.get(str(chat_id), [])
    balance = get_balance(chat_id)
    
    # Автоматически добавляем заработанный баланс к основному
    earned_amount = len(user_referrals) * 500
    if earned_amount > 0:
        # Проверяем, был ли уже добавлен заработок (сохраняем флаг в user_info)
        if not user_info.get('earned_added', False):
            add_balance(chat_id, earned_amount)
            user_info['earned_added'] = True
            users[str(chat_id)] = user_info
            save_users(users)
            balance = get_balance(chat_id)  # Обновляем баланс после добавления
    
    member_date = user_info.get('registered', 'N/A')
    if member_date != 'N/A':
        try:
            member_date = datetime.fromisoformat(member_date).strftime('%d.%m.%Y')
        except:
            pass
    
    lang = load_user_language(chat_id)
    
    if lang == 'en':
        text = (
            f"{get_text(chat_id, 'profile_title')}\n\n"
            f"<b>👤 {user_info.get('first_name', 'User')}</b>\n"
            f"🆔 ID: <code>{chat_id}</code>\n\n"
            f"<b>💰 Balance:</b> <b>{balance}₽</b>\n"
            f"<b>👥 Referrals Invited:</b> <b>{len(user_referrals)}</b>\n"
            f"<b>📅 Member Since:</b> <b>{member_date}</b>\n"
        )
    else:
        text = (
            f"{get_text(chat_id, 'profile_title')}\n\n"
            f"<b>👤 {user_info.get('first_name', 'User')}</b>\n"
            f"🆔 ID: <code>{chat_id}</code>\n\n"
            f"<b>💰 Баланс:</b> <b>{balance}₽</b>\n"
            f"<b>👥 Рефералов приглашено:</b> <b>{len(user_referrals)}</b>\n"
            f"<b>📅 Участник с:</b> <b>{member_date}</b>\n"
        )
    
    msg = bot.send_message(chat_id, text, reply_markup=profile_kbd(chat_id), parse_mode='HTML')
    track_message(chat_id, msg.message_id)


# ==================== ОНБОРДИНГ СИСТЕМА ====================

def send_onboarding_message_1(chat_id):
    """Первое сообщение онбординга - приветствие от меня"""
    text = ("👋 Привет! На связи Я 🤝\n\n"
            "Напиши свой игровой ID, чтобы я мог помочь тебе "
            "со всеми твоими вопросами и помочь с доступом 🎮")
     
    bot.send_message(chat_id, text, parse_mode='HTML')
    set_init_state(chat_id, 'waiting_id')


def send_onboarding_message_2(chat_id):
    """Второе сообщение - ссылка на скачивание"""
    download_text = ("📱 <b>Скачайте приложение:</b>\n"
                    "Для лучшего опыта скачайте наше мобильное приложение.\n\n"
                    "Перейти по ссылке: " + SUBMIT_LINK)
    
    bot.send_message(chat_id, download_text, parse_mode='HTML')


def send_onboarding_message_3(chat_id):
    """Третье сообщение - информация о бонусе"""
    bonus_text = ("🎁 <b>Чтобы забрать бонус:</b>\n\n"
                 "✅ Подтверди свой игровой ID\n"
                 "✅ Сделай первый депозит\n"
                 "✅ Получи <b>100% бонус до 50,000₽</b>\n"
                 "✅ Плюс билет на турнир!\n\n"
                 "<i>Бонус сгорит уже завтра, поэтому поспеши! ⏰</i>")
    
    bot.send_message(chat_id, bonus_text, parse_mode='HTML')
    # Устанавливаем первый таймер напоминания через 20 минут
    schedule_reminder(chat_id, 1, 20)


def send_onboarding_reminder(chat_id, level):
    """Отправляет напоминание по уровню"""
    reminders = {
        1: ("⚙️ Всё получилось? 🤔\n\n"
            "Не получилось скачать приложение? "
            "Напиши и я помогу!"),
        
        2: ("🙋‍♂️ Приложение скачал? 📲\n\n"
            "Пока ты не подтвердишь свой ID, "
            "я не смогу активировать твой бонус 💰"),
        
        3: ("Ты уже почти в игре! 🎮\n\n"
            "Осталось совсем чуть-чуть:\n"
            "1. Скачай приложение\n"
            "2. Подтверди свой игровой ID\n"
            "3. Забери свой бонус 🎁"),
        
        4: ("⏳ Без твоего ID заявку я не могу обработать!\n\n"
            "Срок действия бонуса заканчивается скоро.\n"
            "Напиши просто число без пробелов 👇"),
        
        5: ("🤝 Пригласи друзей и зарабатывай!\n\n"
            "За каждого приглашенного друга ты получишь\n"
            "<b>💰 10% от его бай-ина</b>\n\n"
            "Твоя реферальная ссылка уже готова к использованию! 🚀")
    }
    
    if level in reminders:
        text = reminders[level]
        bot.send_message(chat_id, text, parse_mode='HTML')
        
        # Планируем следующее напоминание
        if level < 5:
            schedule_reminder(chat_id, level + 1, 20)


def schedule_reminder(chat_id, level, delay_minutes):
    """Планирует напоминание через указанное время (в минутах)"""
    def reminder_task():
        try:
            # Проверяем, завершена ли инициализация
            current_state = get_init_state(chat_id)
            if current_state == 'waiting_id':
                send_onboarding_reminder(chat_id, level)
        except:
            pass
    
    # Запускаем поток с задержкой
    delay_seconds = delay_minutes * 60
    timer = threading.Timer(delay_seconds, reminder_task)
    timer.daemon = True
    timer.start()


def send_founder_story(chat_id):
    """Отправляет историю основателя"""
    username = load_users().get(str(chat_id), {}).get('username', 'друг')
    
    story = (f"👋 Привет, @{username}! 🤝\n\n"
            "<b>Кто мы?</b>\n"
            "Мы Kingxxx Club,  У нас работает "
            "<b>20+ профессионалов</b> и "
            "<b>1500+ активных игроков</b>.\n\n"
            "<b>Что мы предлагаем?</b>\n"
            "✅ Честную игру 24/7\n"
            "✅ Лучшие турниры\n"
            "✅ Техническую поддержку в любое время\n"
            "✅ Щедрые бонусы\n\n"
            "<b>Для тебя прямо сейчас:</b>\n"
            "🎁 <b>100% бонус до 50,000₽</b> на первый депозит\n"
            "🎟 <b>Билет на турнир</b> (стоимость 5,000₽)\n\n"
            "Всё это стоит <b>55,000₽</b>, но ты получишь за регистрацию! 🚀\n\n"
            "<i>Предложение действует только сегодня!</i> ⏰")
    
    bot.send_message(chat_id, story, parse_mode='HTML')


def send_onboarding_welcome(chat_id):
    """Отправляет экран приветствия после ввода ID"""
    text = ("👋 Привет! Рад приветствовать вас в Kinxxx Club.\n\n"
            "Kingxxx Club — это пространство для честной и комфортной игры.\n"
            "У нас играют только реальные участники — мы внимательно относимся к качеству сообщества и проводим личные собеседования перед вступлением в клуб.\n\n"
            "🔒 Служба безопасности: контролирует честность игры и оперативно реагирует на обращения, поддерживая комфортную атмосферу за столами.")
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('🚀 Я ГОТОВ', callback_data='onboarding_welcome_ready'))
    
    bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=kb)


def send_onboarding_benefits(chat_id):
    """Отправляет экран с преимуществами"""
    text = ("🎁 <b>Преимущества для новых игроков</b>\n\n"
            "• +50% к первому депозиту\n\n"
            "• Компенсация до 20% потерь, если будет подтверждена нечестная игра\n\n"
            "• Доступ к бонусным игровым миссиям с дополнительными наградами")
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('➡️ Далее', callback_data='onboarding_benefits_next'))
    
    bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=kb)


def send_onboarding_final(chat_id):
    """Отправляет финальный экран с бонусами"""
    text = ("⏳ Бонусы действуют сегодня.\n"
            "Если готовы продолжить — нажмите кнопку «Клуб»")
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('🏆 Клуб', callback_data='onboarding_final_club'))
    
    bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=kb)


def complete_onboarding(chat_id, game_id):
    """Завершает онбординг и показывает новые экраны"""
    set_init_state(chat_id, 'completed')
    
    # Сохраняем игровой ID
    users = load_users()
    user_str = str(chat_id)
    if user_str in users:
        users[user_str]['game_id'] = game_id
        users[user_str]['bonus_claimed'] = False
        save_users(users)
    
    # Показываем экраны онбординга
    bot.send_message(chat_id, "✅ ID сохранен!")
    time.sleep(1)
    send_onboarding_welcome(chat_id)


@bot.message_handler(commands=['start'])
def cmd_start(message):
    chat_id = message.chat.id
    args = message.text.split()
    
    # Загружаем или создаем данные пользователя
    users = load_users()
    user_id_str = str(chat_id)
    
    is_new_user = user_id_str not in users
    
    if is_new_user:
        users[user_id_str] = {
            'username': message.from_user.username or 'unknown',
            'first_name': message.from_user.first_name or 'User',
            'registered': datetime.now().isoformat(),
            'referrer_id': None,
            'game_id': None,
            'bonus_claimed': False
        }
        save_users(users)
    
    # Обработка реферальной ссылки
    if len(args) > 1 and args[1].startswith('ref_'):
        try:
            referrer_id = args[1].replace('ref_', '')
            if referrer_id != user_id_str and referrer_id in users:
                users[user_id_str]['referrer_id'] = referrer_id
                save_users(users)
                
                # Добавляем в реферальную систему
                referrals = load_referrals()
                if referrer_id not in referrals:
                    referrals[referrer_id] = []
                if user_id_str not in referrals[referrer_id]:
                    referrals[referrer_id].append(user_id_str)
                save_referrals(referrals)
                
                bot.send_message(chat_id, 
                    '🎉 Спасибо за регистрацию через реферальную ссылку!\n'
                    '💝 Вы получите специальный бонус')
        except:
            pass
    
    # Если новый пользователь - запускаем онбординг
    if is_new_user:
        # Отправляем серию приветственных сообщений
        send_onboarding_message_1(chat_id)
        
        # Задержка перед вторым сообщением (1 секунда)
        def send_msg2():
            time.sleep(1)
            send_onboarding_message_2(chat_id)
        
        thread2 = threading.Thread(target=send_msg2)
        thread2.daemon = True
        thread2.start()
        
        # Задержка перед третьим сообщением (2 секунды)
        def send_msg3():
            time.sleep(2)
            send_onboarding_message_3(chat_id)
        
        thread3 = threading.Thread(target=send_msg3)
        thread3.daemon = True
        thread3.start()
    else:
        # Для старых пользователей просто показываем меню
        send_menu(chat_id)



@bot.message_handler(commands=['menu'])
def cmd_menu(message):
    send_menu(message.chat.id)


@bot.message_handler(commands=['admin'])
def cmd_admin(message):
    admin_id = message.chat.id
    
    if not is_admin(admin_id):
        bot.send_message(admin_id, '❌ Доступ запрещен!')
        return
    
    # Проверяем тип админа
    if is_support_admin(admin_id):
        text = (
            '⚙️ <b>АДМИН ПАНЕЛЬ ПОДДЕРЖКИ</b> ⚙️\n\n'
            '🔐 <b>Добро пожаловать, администратор поддержки!</b>\n\n'
            '📋 Доступные функции:\n'
            '  • 📞 Настройки поддержки\n\n'
            '⬇️ Выберите действие ниже:'
        )
    else:
        text = (
            '⚙️ <b>АДМИН ПАНЕЛЬ</b> ⚙️\n\n'
            '🔐 <b>Добро пожаловать, администратор!</b>\n\n'
            '📋 Доступные функции:\n'
            '  • 💰 Управление балансом\n'
            '  • 👥 Управление пользователями\n'
            '  • 📋 Просмотр всех пользователей\n'
            '  • 📊 Статистика системы\n'
            '  • 🔍 Изучение пользователей\n'
            '  • 📢 Рассылка сообщений всем\n'
            '  • 🔑 Управление админами\n'
            '  • 📞 Настройки поддержки\n'
            '  • 📥 Скачать информацию\n\n'
            '⬇️ Выберите действие ниже:'
        )
    
    bot.send_message(admin_id, text, reply_markup=admin_menu_kbd(admin_id), parse_mode='HTML')


@bot.message_handler(commands=['instruction'])
def cmd_instruction(message):
    """Инструкция по установке покер-клиентов"""
    chat_id = message.chat.id
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('🎮 ClubGG', callback_data='instruction_clubgg'))
    kb.add(types.InlineKeyboardButton('🎮 PPPoker', callback_data='instruction_pppoker'))
    
    text = (
        "📚 <b>ВЫБЕРИТЕ КЛИЕНТ ДЛЯ УСТАНОВКИ</b>\n\n"
        "Здесь вы найдете пошаговую инструкцию по установке и регистрации "
        "в одном из популярных покер-клиентов:"
    )
    
    bot.send_message(chat_id, text, reply_markup=kb, parse_mode='HTML')


@bot.callback_query_handler(func=lambda call: call.data == 'instruction_clubgg')
def callbacks_instruction_clubgg(call):
    """Инструкция по установке ClubGG"""
    chat_id = call.message.chat.id
    
    # Шаг 1: Ссылка на установку
    text1 = (
        "🎮 <b>УСТАНОВКА ClubGG</b>\n\n"
        "<b>Шаг 1: Скачивание и установка</b>\n\n"
        "Тапните по ссылке ниже, чтобы скачать ClubGG:"
    )
    
    kb1 = types.InlineKeyboardMarkup()
    kb1.add(types.InlineKeyboardButton('📥 Скачать ClubGG', url='https://clubgg.app.link/bbCYbP4wWXb'))
    kb1.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='instruction_back'))
    
    try:
        bot.edit_message_text(text1, chat_id, call.message.message_id, 
                            reply_markup=kb1, parse_mode='HTML')
    except:
        bot.send_message(chat_id, text1, reply_markup=kb1, parse_mode='HTML')
    
    # Шаг 2: Фото регистрации
    try:
        with open('registration.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 2: Регистрация в ClubGG</b>\n\nСледуйте этим инструкциям:",
                parse_mode='HTML'
            )
    except:
        bot.send_message(chat_id, "<b>Шаг 2: Регистрация в ClubGG</b>\n\nФайл registration.png не найден", parse_mode='HTML')
    
    # Шаг 3: Поиск айди
    try:
        with open('findid1.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 3: Поиск своего ID (часть 1)</b>",
                parse_mode='HTML'
            )
    except:
        pass
    
    try:
        with open('findid2.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 3: Поиск своего ID (часть 2)</b>\n\n✅ Ваш ID найден! Используйте его для пополнения баланса.",
                parse_mode='HTML'
            )
    except:
        pass
    
    # Кнопка для возврата
    kb_final = types.InlineKeyboardMarkup()
    kb_final.add(types.InlineKeyboardButton('💳 Перейти на оплату', url='https://t.me/sapayobot?start=PokerKingClubBot_game_id'))
    kb_final.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='instruction_back'))
    
    bot.send_message(
        chat_id,
        "✅ <b>Инструкция по установке ClubGG завершена!</b>\n\n"
        "Теперь вы можете пополнить баланс и начать играть 🎮",
        reply_markup=kb_final,
        parse_mode='HTML'
    )
    
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'instruction_pppoker')
def callbacks_instruction_pppoker(call):
    """Инструкция по установке PPPoker"""
    chat_id = call.message.chat.id
    
    # Шаг 1: Скачивание
    text1 = (
        "🎮 <b>УСТАНОВКА PPPoker</b>\n\n"
        "<b>Шаг 1: Скачивание приложения</b>\n\n"
        "Смотрите инструкцию ниже:"
    )
    
    kb1 = types.InlineKeyboardMarkup()
    kb1.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='instruction_back'))
    
    try:
        bot.edit_message_text(text1, chat_id, call.message.message_id, 
                            reply_markup=kb1, parse_mode='HTML')
    except:
        bot.send_message(chat_id, text1, reply_markup=kb1, parse_mode='HTML')
    
    # Шаг 1: Фото скачивания
    try:
        with open('downladpppoker.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 1: Скачивание PPPoker</b>\n\nСкачайте приложение по этой инструкции",
                parse_mode='HTML'
            )
    except:
        bot.send_message(chat_id, "<b>Шаг 1: Скачивание PPPoker</b>\n\nФайл downladpppoker.png не найден", parse_mode='HTML')
    
    # Шаг 2: Регистрация
    try:
        with open('pppokerreg.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 2: Регистрация в PPPoker</b>\n\nЗарегистрируйтесь по этой инструкции",
                parse_mode='HTML'
            )
    except:
        pass
    
    # Шаг 3: Вступление в клуб (часть 1)
    try:
        with open('joinclub1.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 3: Вступление в клуб (часть 1)</b>",
                parse_mode='HTML'
            )
    except:
        pass
    
    # Шаг 3: Вступление в клуб (часть 2) и ввод кода
    try:
        with open('joinclub2.png', 'rb') as photo:
            bot.send_photo(
                chat_id,
                photo,
                caption="<b>Шаг 3: Вступление в клуб (часть 2)</b>\n\n"
                        "Введите код клуба: <code>4728345</code>\n\n"
                        "После нажатия кнопки вы присоединитесь к клубу через несколько минут ✅",
                parse_mode='HTML'
            )
    except:
        bot.send_message(
            chat_id,
            "<b>Вступление в клуб</b>\n\n"
            "Введите код клуба: <code>4728345</code>\n\n"
            "После нажатия вы присоединитесь к клубу через несколько минут ✅",
            parse_mode='HTML'
        )
    
    # Кнопка для возврата
    kb_final = types.InlineKeyboardMarkup()
    kb_final.add(types.InlineKeyboardButton('💳 Перейти на оплату', url='https://t.me/sapayobot?start=PokerKingClubBot_game_id'))
    kb_final.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='instruction_back'))
    
    bot.send_message(
        chat_id,
        "✅ <b>Инструкция по установке PPPoker завершена!</b>\n\n"
        "Теперь вы можете пополнить баланс и начать играть 🎮",
        reply_markup=kb_final,
        parse_mode='HTML'
    )
    
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'instruction_back')
def callbacks_instruction_back(call):
    """Возврат в меню инструкций"""
    bot.delete_message(call.message.chat.id, call.message.message_id)
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('🎮 ClubGG', callback_data='instruction_clubgg'))
    kb.add(types.InlineKeyboardButton('🎮 PPPoker', callback_data='instruction_pppoker'))
    
    text = (
        "📚 <b>ВЫБЕРИТЕ КЛИЕНТ ДЛЯ УСТАНОВКИ</b>\n\n"
        "Здесь вы найдете пошаговую инструкцию по установке и регистрации "
        "в одном из популярных покер-клиентов:"
    )
    
    bot.send_message(call.message.chat.id, text, reply_markup=kb, parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'promos')
def callbacks_promos(call):
    track_action(call.message.chat.id, 'promos')  # Отслеживаем действие
    lang = load_user_language(call.message.chat.id)
    
    if lang == 'en':
        text = (
            '💎 <b>EXCLUSIVE BONUSES </b> \n\n'
            '<b>🔥 BAD BEAT PROTECTION</b>\n'
            '   Your insurance when luck runs out\n\n'
            '<b>🎰 FREE TOURNAMENTS</b>\n'
            '   Play for free, win real money\n\n'
            '<b>💵 REFERRAL PROGRAM</b>\n'
            '   Build your passive income stream\n\n'
            '<i>👇 Select a bonus to learn more</i>'
        )
    else:
        text = (
            '💎 <b>ЭКСКЛЮЗИВНЫЕ БОНУСЫ  </b> \n\n'
            '<b>🔥 ЗАЩИТА ОТ БАД БИТОВ</b>\n'
            '   Твоя страховка при невезении\n\n'
            '<b>🎰 БЕСПЛАТНЫЕ ТУРНИРЫ</b>\n'
            '   Играй бесплатно, выигрывай реально\n\n'
            '<b>💵 РЕФЕРАЛЬНАЯ ПРОГРАММА</b>\n'
            '   Создай поток пассивного дохода\n\n'
            '<i>👇 Выбери бонус для подробности</i>'
        )
    
    try:
        # Пытаемся удалить старое сообщение и отправить новое с картинкой
        try:
            bot.delete_message(call.message.chat.id, call.message.message_id)
        except:
            pass
        
        # Отправляем картинку с текстом и кнопками
        with open('акции.png', 'rb') as photo:
            msg = bot.send_photo(call.message.chat.id, photo, caption=text, reply_markup=promos_kbd(call.message.chat.id), parse_mode='HTML')
            track_message(call.message.chat.id, msg.message_id)
    except Exception as e:
        # Fallback: если картинки нет, отправляем обычное сообщение
        try:
            bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=promos_kbd(call.message.chat.id), parse_mode='HTML')
        except:
            msg = bot.send_message(call.message.chat.id, text, reply_markup=promos_kbd(call.message.chat.id), parse_mode='HTML')
            track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'payment_info')
def callbacks_payment_info(call):
    """Показывает информацию о платеже"""
    track_action(call.message.chat.id, 'payment')
    user_id = call.message.chat.id
    
    users = load_users()
    user_str = str(user_id)
    user_info = users.get(user_str, {})
    
    referrer_id = user_info.get('referrer_id')
    game_id = user_info.get('game_id')
    
    # Если нет game_id - просим ввести
    if not game_id:
        user_states[user_id] = 'waiting_payment_game_id'
        bot.send_message(
            user_id,
            "📋 <b>Для оплаты нужен ваш игровой ID</b>\n\n"
            "Пожалуйста, отправьте ваш игровой ID:",
            parse_mode='HTML'
        )
        bot.answer_callback_query(call.id)
        return
    
    # Создаем ссылку с игровым ID
    payment_url = f'https://t.me/sapayobot?start=PokerKingClubBot_{game_id}'
    
    # Если нет реферера - показываем обычный платеж
    if not referrer_id or referrer_id not in users:
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('💳 Перейти на оплату', url=payment_url))
        kb.add(types.InlineKeyboardButton('🎁 Пополнить промо-кодом', callback_data='use_promo_code'))
        kb.add(types.InlineKeyboardButton('🎙️ Сообщество и акции', url=get_setting('COMMUNITY_LINK', COMMUNITY_LINK)))
        kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='menu'))
        
        text = (
            "💳 <b>ОПЛАТА</b>\n\n"
            f"🎮 Ваш ID: <code>{game_id}</code>\n\n"
            "<b>Способы пополнения:</b>\n\n"
            "1️⃣ <b>💳 Перейти на оплату</b> - Быстрый платеж через внешний сервис\n\n"
            "2️⃣ <b>🎁 Пополнить промо-кодом</b> - Используйте промо-коды для бесплатного пополнения\n\n"
            "💡 <b>Совет:</b> Вступите в наше сообщество \"🎙️ Сообщество и акции\" чтобы получить промо-коды и участвовать в акциях! 🎉"
        )
    else:
        # Показываем платеж с информацией реферера
        referrer_info = users.get(str(referrer_id), {})
        referrer_name = referrer_info.get('first_name', 'Администратор')
        referrer_username = referrer_info.get('username', 'N/A')
        
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('💳 Перейти на оплату', url=payment_url))
        kb.add(types.InlineKeyboardButton('🎁 Пополнить промо-кодом', callback_data='use_promo_code'))
        kb.add(types.InlineKeyboardButton('🎙️ Сообщество и акции', url=get_setting('COMMUNITY_LINK', COMMUNITY_LINK)))
        kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='menu'))
        
        text = (
            "💳 <b>ОПЛАТА</b>\n\n"
            f"🎮 Ваш ID: <code>{game_id}</code>\n\n"
            f"<b>Реферер:</b> <code>@{referrer_username}</code>\n"
            f"({referrer_name})\n\n"
            "<b>Способы пополнения:</b>\n\n"
            "1️⃣ <b>💳 Перейти на оплату</b> - Быстрый платеж через внешний сервис\n\n"
            "2️⃣ <b>🎁 Пополнить промо-кодом</b> - Используйте промо-коды для бесплатного пополнения\n\n"
            "💡 <b>Совет:</b> Вступите в наше сообщество \"🎙️ Сообщество и акции\" чтобы получить промо-коды и участвовать в акциях! 🎉"
        )
    
    try:
        bot.edit_message_text(text, user_id, call.message.message_id, 
                            reply_markup=kb, parse_mode='HTML')
    except:
        bot.send_message(user_id, text, reply_markup=kb, parse_mode='HTML')
    
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'waiting_payment_game_id')
def process_payment_game_id(message):
    """Обработка ввода game_id для платежа"""
    user_id = message.chat.id
    game_id = message.text.strip()
    
    # Проверяем, что ID содержит только цифры
    if not game_id or not game_id.isalnum():
        bot.send_message(user_id, "❌ ID должен содержать только цифры и буквы. Попробуйте снова!")
        return
    
    # Сохраняем game_id
    users = load_users()
    user_str = str(user_id)
    if user_str in users:
        users[user_str]['game_id'] = game_id
        save_users(users)
    
    user_states.pop(user_id, None)
    
    # Создаем ссылку с игровым ID
    payment_url = f'https://t.me/sapayobot?start={game_id}'
    
    # Теперь показываем информацию о платеже
    users = load_users()
    user_info = users.get(user_str, {})
    referrer_id = user_info.get('referrer_id')
    
    # Если нет реферера
    if not referrer_id or referrer_id not in users:
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('💳 Перейти на оплату', url=payment_url))
        kb.add(types.InlineKeyboardButton('⬅️ Назад в меню', callback_data='menu'))
        
        text = (
            "✅ <b>ID сохранен!</b>\n\n"
            "💳 <b>ОПЛАТА</b>\n\n"
            f"🎮 Ваш ID: <code>{game_id}</code>\n\n"
            "Нажмите кнопку ниже для перехода на страницу оплаты"
        )
    else:
        # Показываем платеж с информацией реферера
        referrer_info = users.get(str(referrer_id), {})
        referrer_name = referrer_info.get('first_name', 'Администратор')
        referrer_username = referrer_info.get('username', 'N/A')
        
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('💳 Перейти на оплату', url=payment_url))
        kb.add(types.InlineKeyboardButton('⬅️ Назад в меню', callback_data='menu'))
        
        text = (
            "✅ <b>ID сохранен!</b>\n\n"
            "💳 <b>ОПЛАТА</b>\n\n"
            f"🎮 Ваш ID: <code>{game_id}</code>\n\n"
            f"<b>Реферер:</b> <code>@{referrer_username}</code>\n"
            f"({referrer_name})\n\n"
            "Нажмите кнопку ниже для перехода на страницу оплаты"
        )
    
    bot.send_message(user_id, text, reply_markup=kb, parse_mode='HTML')


@bot.callback_query_handler(func=lambda call: call.data == 'profile')
def callbacks_profile(call):
    track_action(call.message.chat.id, 'profile')
    send_profile(call.message.chat.id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'menu')
def callbacks_menu(call):
    bot.answer_callback_query(call.id)
    send_menu(call.message.chat.id)


@bot.callback_query_handler(func=lambda call: call.data == 'language')
def callbacks_language(call):
    track_action(call.message.chat.id, 'language')
    text = 'Выберите язык / Select Language:'
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=language_kbd())
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=language_kbd())
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data in ['lang_ru', 'lang_en'])
def callbacks_change_language(call):
    lang = call.data.replace('lang_', '')
    save_user_language(call.message.chat.id, lang)
    
    if lang == 'ru':
        text = '✅ Язык изменен на Русский'
    else:
        text = '✅ Language changed to English'
    
    bot.answer_callback_query(call.id, text, show_alert=True)
    # Автоматически возвращаемся в меню после смены языка
    send_menu(call.message.chat.id)


@bot.callback_query_handler(func=lambda call: call.data == 'community_menu')
def callbacks_community_menu(call):
    track_action(call.message.chat.id, 'community')
    lang = load_user_language(call.message.chat.id)
    
    if lang == 'en':
        text = '<b>👥 COMMUNITY MENU</b>\n\nChoose an option:'
    else:
        text = '<b>👥 МЕНЮ СООБЩЕСТВА</b>\n\nВыберите опцию:'
    
    # Отправляем картинку если она доступна
    try:
        if MENU_IMAGE and (MENU_IMAGE.startswith('http') or os.path.exists(MENU_IMAGE)):
            if MENU_IMAGE.startswith('http'):
                bot.edit_message_media(
                    media=types.InputMediaPhoto(media=MENU_IMAGE, caption=text, parse_mode='HTML'),
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id
                )
                bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=community_menu_kbd(call.message.chat.id))
            else:
                with open(MENU_IMAGE, 'rb') as photo:
                    bot.edit_message_media(
                        media=types.InputMediaPhoto(media=photo, caption=text, parse_mode='HTML'),
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id
                    )
                    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=community_menu_kbd(call.message.chat.id))
        else:
            bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=community_menu_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        try:
            # Если edit не сработал, отправляем новое сообщение
            if MENU_IMAGE and (MENU_IMAGE.startswith('http') or os.path.exists(MENU_IMAGE)):
                if MENU_IMAGE.startswith('http'):
                    msg = bot.send_photo(call.message.chat.id, photo=MENU_IMAGE, caption=text, reply_markup=community_menu_kbd(call.message.chat.id), parse_mode='HTML')
                else:
                    with open(MENU_IMAGE, 'rb') as photo:
                        msg = bot.send_photo(call.message.chat.id, photo=photo, caption=text, reply_markup=community_menu_kbd(call.message.chat.id), parse_mode='HTML')
                track_message(call.message.chat.id, msg.message_id)
            else:
                msg = bot.send_message(call.message.chat.id, text, reply_markup=community_menu_kbd(call.message.chat.id), parse_mode='HTML')
                track_message(call.message.chat.id, msg.message_id)
        except:
            msg = bot.send_message(call.message.chat.id, text, reply_markup=community_menu_kbd(call.message.chat.id), parse_mode='HTML')
            track_message(call.message.chat.id, msg.message_id)
    
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'apply_conditions')
def callbacks_apply_conditions(call):
    track_action(call.message.chat.id, 'apply')
    lang = load_user_language(call.message.chat.id)
    
    if lang == 'en':
        text = '<b>✅ SUBMIT APPLICATION</b>\n\nChoose an option:'
    else:
        text = '<b>✅ ПОДАТЬ ЗАЯВКУ</b>\n\nВыберите опцию:'
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=apply_conditions_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, text, reply_markup=apply_conditions_kbd(call.message.chat.id), parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'submit_application')
def callbacks_submit_application(call):
    """Открывает форму подачи заявки"""
    user_id = call.message.chat.id
    track_action(user_id, 'submit_app')
    
    users = load_users()
    user_info = users.get(str(user_id), {})
    game_id = user_info.get('game_id')
    
    # Если нет game_id - просим ввести
    if not game_id:
        user_states[user_id] = 'waiting_application_game_id'
        bot.send_message(
            user_id,
            "📋 <b>Для подачи заявки нужен ваш игровой ID</b>\n\n"
            "Пожалуйста, отправьте ваш игровой ID:",
            parse_mode='HTML'
        )
        bot.answer_callback_query(call.id)
        return
    
    # Отправляем заявку админам
    applications = load_applications()
    app_id = str(int(time.time()))
    
    username = user_info.get('username', 'unknown')
    
    applications[app_id] = {
        'client_id': user_id,
        'client_name': user_info.get('first_name', 'Клиент'),
        'username': username,
        'game_id': game_id,
        'status': 'pending',
        'created_at': datetime.now().isoformat()
    }
    save_applications(applications)
    
    # Загружаем админов
    admins = load_admins()
    
    # Фильтруем админов, у которых включены заявки
    active_admins = [int(admin_id) for admin_id in admins.keys() if is_applications_enabled(int(admin_id))]
    
    if active_admins:
        # Отправляем уведомление админам
        notification_text = (
            f"📋 <b>НОВАЯ ЗАЯВКА</b>\n\n"
            f"👤 Клиент: <b>{user_info.get('first_name', 'Клиент')}</b>\n"
            f"📱 Username: <code>@{username}</code>\n"
            f"🎮 Игровой ID: <code>{game_id}</code>\n\n"
            f"🆔 ID заявки: <code>{app_id}</code>"
        )
        
        for admin_id in active_admins:
            kb = types.InlineKeyboardMarkup()
            kb.add(types.InlineKeyboardButton('✅ Добавлен', callback_data=f'app_approve_{app_id}'))
            kb.add(types.InlineKeyboardButton('❌ Отклонить', callback_data=f'app_reject_{app_id}'))
            
            try:
                bot.send_message(admin_id, notification_text, reply_markup=kb, parse_mode='HTML')
            except:
                pass
        
        bot.send_message(user_id, "✅ Ваша заявка отправлена администраторам!\n\nОжидайте ответа...", parse_mode='HTML')
    else:
        bot.send_message(user_id, "⏳ Сейчас нет доступных администраторов. Попробуйте позже.", parse_mode='HTML')
    
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'waiting_application_game_id')
def process_application_game_id(message):
    """Обработка ввода game_id для заявки"""
    user_id = message.chat.id
    game_id = message.text.strip()
    
    # Проверяем, что ID содержит только цифры
    if not game_id or not game_id.isalnum():
        bot.send_message(user_id, "❌ ID должен содержать только цифры и буквы. Попробуйте снова!")
        return
    
    # Сохраняем game_id
    users = load_users()
    user_str = str(user_id)
    if user_str in users:
        users[user_str]['game_id'] = game_id
        save_users(users)
    
    user_states.pop(user_id, None)
    
    # Отправляем заявку админам
    applications = load_applications()
    app_id = str(int(time.time()))
    
    user_info = users.get(user_str, {})
    username = user_info.get('username', 'unknown')
    
    applications[app_id] = {
        'client_id': user_id,
        'client_name': user_info.get('first_name', 'Клиент'),
        'username': username,
        'game_id': game_id,
        'status': 'pending',
        'created_at': datetime.now().isoformat()
    }
    save_applications(applications)
    
    # Загружаем админов
    admins = load_admins()
    
    # Фильтруем админов, у которых включены заявки
    active_admins = [int(admin_id) for admin_id in admins.keys() if is_applications_enabled(int(admin_id))]
    
    if active_admins:
        # Отправляем уведомление админам
        notification_text = (
            f"📋 <b>НОВАЯ ЗАЯВКА</b>\n\n"
            f"👤 Клиент: <b>{user_info.get('first_name', 'Клиент')}</b>\n"
            f"📱 Username: <code>@{username}</code>\n"
            f"🎮 Игровой ID: <code>{game_id}</code>\n\n"
            f"🆔 ID заявки: <code>{app_id}</code>"
        )
        
        for admin_id in active_admins:
            kb = types.InlineKeyboardMarkup()
            kb.add(types.InlineKeyboardButton('✅ Добавлен', callback_data=f'app_approve_{app_id}'))
            kb.add(types.InlineKeyboardButton('❌ Отклонить', callback_data=f'app_reject_{app_id}'))
            
            try:
                bot.send_message(admin_id, notification_text, reply_markup=kb, parse_mode='HTML')
            except:
                pass
        
        bot.send_message(user_id, "✅ ID сохранен! Ваша заявка отправлена администраторам!\n\nОжидайте ответа...", parse_mode='HTML')
    else:
        bot.send_message(user_id, "⏳ Сейчас нет доступных администраторов. Попробуйте позже.", parse_mode='HTML')


@bot.callback_query_handler(func=lambda call: call.data.startswith('app_approve_') or call.data.startswith('app_reject_'))
def callbacks_application_decision(call):
    """Админ одобрил или отклонил заявку"""
    admin_id = call.message.chat.id
    
    if not is_admin(admin_id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    # Проверяем что это админ поддержки или полный админ
    if not is_support_admin(admin_id) and not is_full_admin(admin_id):
        bot.answer_callback_query(call.id, '❌ Только администраторы могут это делать!', show_alert=True)
        return
    
    if call.data.startswith('app_approve_'):
        app_id = call.data.replace('app_approve_', '')
        action = 'approve'
    else:
        app_id = call.data.replace('app_reject_', '')
        action = 'reject'
    
    applications = load_applications()
    
    if app_id not in applications:
        bot.answer_callback_query(call.id, '❌ Заявка не найдена', show_alert=True)
        return
    
    app = applications[app_id]
    client_id = app['client_id']
    client_name = app['client_name']
    admin_name = call.message.chat.first_name or 'Администратор'
    
    if action == 'approve':
        applications[app_id]['status'] = 'approved'
        message_text = f"✅ <b>Вы добавлены в Kinxxx Club!</b>\n\nАдминистратор {admin_name} подтвердил вашу заявку. Добро пожаловать! 🎉"
        admin_response = f"✅ Вы одобрили заявку пользователя {client_name}"
    else:
        applications[app_id]['status'] = 'rejected'
        message_text = f"❌ <b>Ваша заявка отклонена</b>\n\nАдминистратор {admin_name} отклонил вашу заявку. Попробуйте позже или свяжитесь с поддержкой."
        admin_response = f"❌ Вы отклонили заявку пользователя {client_name}"
    
    save_applications(applications)
    
    # Отправляем уведомление клиенту
    bot.send_message(client_id, message_text, parse_mode='HTML')
    
    # Уведомляем админа
    bot.edit_message_text(
        f"📋 <b>ЗАЯВКА</b>\n\n"
        f"👤 Клиент: <b>{client_name}</b>\n"
        f"📱 Username: <code>@{app['username']}</code>\n"
        f"🎮 Игровой ID: <code>{app['game_id']}</code>\n\n"
        f"{admin_response}",
        admin_id, call.message.message_id, parse_mode='HTML'
    )
    
    bot.answer_callback_query(call.id, admin_response, show_alert=False)


@bot.callback_query_handler(func=lambda call: call.data == 'show_conditions')
def callbacks_show_conditions(call):
    lang = load_user_language(call.message.chat.id)
    
    conditions_text = (
        '<b>⚖️ ОБЩИЕ УСЛОВИЯ</b>\n\n'
        '• Бонусы доступны только новым игрокам.\n'
        '• Клуб может отменить бонусы при нарушении правил.\n'
        '• Каждый бонус выдаётся один раз.\n'
        '• Все начисления контролируются службой безопасности.\n'
        '• Поддержка работает 24/7.\n\n'
        '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n'
        '<b>🎁 БОНУСЫ</b>\n\n'
        '<b>+50% к первому депозиту</b>\n'
        '— Бонус начисляется после первого пополнения.\n'
        '— Бонус недоступен для вывода.\n'
        '— Минимальная игровая активность для вывода выигрыша: 3000 раздач.\n\n'
        '<b>Компенсация до 20% потерь при нечестной игре</b>\n'
        '— Выдаётся после подтверждения нарушения службой безопасности.\n\n'
        '<b>Розыгрыши ценных призов для новых игроков</b>\n'
        '— Участие доступно при активном выполнении задач клуба.'
    )
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton(get_text(call.message.chat.id, 'back'), callback_data='apply_conditions'))
    
    try:
        bot.edit_message_text(conditions_text, call.message.chat.id, call.message.message_id, reply_markup=kb, parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, conditions_text, reply_markup=kb, parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'community_menu_back')
def callbacks_community_menu_back(call):
    callbacks_community_menu(call)


@bot.callback_query_handler(func=lambda call: call.data == 'apply')
def callbacks_apply(call):
    track_action(call.message.chat.id, 'apply')
    lang = load_user_language(call.message.chat.id)
    
    if lang == 'en':
        text = '📋 <b>APPLICATION SUBMITTED</b>\n\nOur manager will contact you shortly!\n\n👉 Or write directly to @karinakroks'
    else:
        text = '📋 <b>ЗАЯВКА ОТПРАВЛЕНА</b>\n\nНаш менеджер вскоре с вами свяжется!\n\n👉 Или напиши напрямую @karinakroks'
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=back_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, text, reply_markup=back_kbd(call.message.chat.id), parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'bonus')
def callbacks_bonus(call):
    track_action(call.message.chat.id, 'bonus')
    lang = load_user_language(call.message.chat.id)
    
    if lang == 'en':
        text = (
            '<b>🔥 BAD BEAT PROTECTION</b>\n'
            '───────────────────\n'
            '✓ Applies to losses > 4 of a kind\n'
            '✓ Automatic payment without documentation\n'
            '✓ Withdraw anytime\n\n'
            '<b>🎰 FREE TOURNAMENT ENTRY</b>\n'
            '───────────────────\n'
            '✓ No deposit required\n'
            '✓ Real money winnings\n'
            '✓ tournaments available\n\n'
            '<b>💵 REFERRAL REWARDS</b>\n'
            '───────────────────\n'
            '✓ 10% lifetime commission\n'
            '✓ Unlimited referrals\n'
            '✓ Instant payouts\n\n'
            '⏰ <b>All bonuses active NOW!</b>'
        )
    else:
        text = (
            '<b>🔥 ЗАЩИТА ОТ БАД БИТОВ</b>\n'
            '───────────────────\n'
            '✓ Срабатывает при проигрыше > каре\n'
            '✓ Автоматический платеж без документов\n'
            '✓ Выводи когда угодно\n\n'
            '<b>🎰 ВХОД В ТУРНИРЫ БЕСПЛАТНО</b>\n'
            '───────────────────\n'
            '✓ Без необходимости вносить депозит\n'
            '✓ Реальные денежные выигрыши\n'
            '✓ Турниры \n\n'
            '<b>💵 ВОЗНАГРАЖДЕНИЕ ЗА РЕФЕРАЛОВ</b>\n'
            '───────────────────\n'
            '✓ 10% пожизненно от каждого\n'
            '✓ Неограниченно рефералов\n'
            '✓ Выплаты мгновенно\n\n'
            '⏰ <b>Все бонусы активны ПРЯМО СЕЙЧАС!</b>'
        )
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=bonus_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, text, reply_markup=bonus_kbd(call.message.chat.id), parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'referral')
def callbacks_referral(call):
    track_action(call.message.chat.id, 'referral')
    lang = load_user_language(call.message.chat.id)
    
    if lang == 'en':
        text = (
            '                                         \n'
            '   � <b>REFERRAL PROGRAM</b> �         \n'
            '                                         \n\n'
            '<b>💡 HOW DOES IT WORK?</b>\n\n'
            '1️⃣ Get your personal link\n'
            '2️⃣ Invite friends via this link\n'
            '3️⃣ Get commission for each one\n\n'
            '<b>💰 REWARD SIZE</b>\n'
            '   ➜ <b>10%</b> from friend\'s first buy-in\n\n'
            '<b>⭐ BENEFITS</b>\n'
            '   ✅ Unlimited income\n'
            '   ✅ Quick payouts\n'
            '   ✅ No fees\n\n'
            'Click the button below for the link!'
        )
    else:
        text = (
            '                                         \n'
            '   👥 <b>РЕФЕРАЛЬНАЯ ПРОГРАММА</b> 👥   \n'
            '                                         \n'
            '<b>💡 КАК ЭТО РАБОТАЕТ?</b>\n\n'
            '1️⃣ Получи персональную ссылку\n'
            '2️⃣ Пригласи друзей по этой ссылке\n'
            '3️⃣ Получай комиссию за каждого\n\n'
            '<b>💰 РАЗМЕР НАГРАДЫ</b>\n'
            '   ➜ <b>10%</b> от первого бай-ина друга\n\n'
            '<b>⭐ ПРЕИМУЩЕСТВА</b>\n'
            '   ✅ Неограниченный доход\n'
            '   ✅ Быстрое начисление\n'
            '   ✅ Никаких комиссий\n\n'
            'Нажми кнопку ниже для ссылки!'
        )
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=referral_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, text, reply_markup=referral_kbd(call.message.chat.id), parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'copy_ref')
def callbacks_copy_ref(call):
    track_action(call.message.chat.id, 'copy_ref')
    lang = load_user_language(call.message.chat.id)
    ref_link = get_referral_link(call.message.chat.id)
    
    if lang == 'en':
        text = (
            '                                         \n'
            '   📤 <b>YOUR REFERRAL LINK</b> 📤      \n'
            '                                         \n\n'
            'Your personal link:\n\n'
            f'<code>{ref_link}</code>\n\n'
            '✅ <b>Link copied to clipboard!</b>\n\n'
            '💡 <b>Tip:</b> Share your link with friends\n'
            'in Telegram or other social networks!'
        )
    else:
        text = (
            '                                         \n'
            '   📤 <b>ТВОЯ РЕФЕРАЛЬНАЯ ССЫЛКА</b> 📤   \n'
            '                                         \n\n'
            'Твоя персональная ссылка:\n\n'
            f'<code>{ref_link}</code>\n\n'
            '✅ <b>Ссылка скопирована в буфер обмена!</b>\n\n'
            '💡 <b>Совет:</b> поделись ссылкой с друзьями\n'
            'в Telegram или других соцсетях!'
        )
    
    kb = types.InlineKeyboardMarkup()
    # Кнопка для отправки ссылки в Избранное (которая копирует в буфер)
    kb.add(types.InlineKeyboardButton('📋 Copy' if lang == 'en' else '📋 Копировать', switch_inline_query_current_chat=ref_link))
    kb.add(types.InlineKeyboardButton('↩️ Back' if lang == 'en' else '↩️ Назад', callback_data='menu'))
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, 
                            reply_markup=kb, parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, text, reply_markup=kb, parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'ref_stats')
def callbacks_ref_stats(call):
    track_action(call.message.chat.id, 'ref_stats')
    lang = load_user_language(call.message.chat.id)
    referrals_data = load_referrals()
    user_referrals = referrals_data.get(str(call.message.chat.id), [])
    
    if lang == 'en':
        text = (
            '                                         \n'
            '   📊 <b>REFERRAL STATISTICS</b> 📊     \n'
            '                                         \n\n'
            f'👥 Total invited: <b>{len(user_referrals)}</b>\n'
            f'💰 Estimated income: <b>{len(user_referrals) * 500}₽</b>\n\n'
            'Invite more friends and\n'
            'get more rewards! 🚀'
        )
    else:
        text = (
            '                                         \n'
            '   📊 <b>СТАТИСТИКА РЕФЕРАЛОВ</b> 📊     \n'
            '                                         \n\n'
            f'👥 Всего приглашено: <b>{len(user_referrals)}</b>\n'
            f'💰 Примерный доход: <b>{len(user_referrals) * 500}₽</b>\n\n'
            'Приглашай больше друзей и\n'
            'получай больше награды! 🚀'
        )
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=referral_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        msg = bot.send_message(call.message.chat.id, text, reply_markup=referral_kbd(call.message.chat.id), parse_mode='HTML')
        track_message(call.message.chat.id, msg.message_id)
    bot.answer_callback_query(call.id)




# Обработчик для ввода игрового ID во время онбординга
@bot.message_handler(func=lambda message: get_init_state(message.chat.id) == 'waiting_id')
def process_game_id(message):
    chat_id = message.chat.id
    game_id = message.text.strip()
    
    # Проверяем, что ID содержит только цифры (или цифры и буквы)
    if not game_id or not game_id.replace(' ', '').isalnum():
        bot.send_message(chat_id, "❌ ID должен содержать только цифры. Попробуй снова!")
        return
    
    # Завершаем онбординг
    complete_onboarding(chat_id, game_id)


@bot.message_handler(commands=['submit'])
def cmd_submit(message):
    user_states[message.chat.id] = 'waiting_submission'
    bot.send_message(message.chat.id, 
        '📝 Заполни форму:\n\n'
        'Отправь свой ник и желаемую сумму бай-ина\n\n'
        'Пример: Nick_123 5000')


# Обработчики для экранов онбординга
@bot.callback_query_handler(func=lambda call: call.data == 'onboarding_welcome_ready')
def callbacks_onboarding_welcome(call):
    chat_id = call.message.chat.id
    bot.edit_message_text("✅ Спасибо!", chat_id, call.message.message_id)
    time.sleep(1)
    send_onboarding_benefits(chat_id)


@bot.callback_query_handler(func=lambda call: call.data == 'onboarding_benefits_next')
def callbacks_onboarding_benefits(call):
    chat_id = call.message.chat.id
    bot.edit_message_text("✅ Отлично!", chat_id, call.message.message_id)
    time.sleep(1)
    send_onboarding_final(chat_id)


@bot.callback_query_handler(func=lambda call: call.data == 'onboarding_final_club')
def callbacks_onboarding_final(call):
    chat_id = call.message.chat.id
    bot.edit_message_text("✅ Добро пожаловать в клуб!", chat_id, call.message.message_id)
    time.sleep(1)
    send_menu(chat_id)


@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'waiting_submission')
def process_submission(message):
    submission = {
        'user_id': message.from_user.id,
        'username': message.from_user.username or 'unknown',
        'first_name': message.from_user.first_name or 'User',
        'text': message.text,
        'timestamp': datetime.now().isoformat()
    }
    # append to submissions file
    out_path = os.path.join(os.path.dirname(__file__), 'submissions.json')
    try:
        if os.path.exists(out_path):
            with open(out_path, 'r', encoding='utf-8') as f:
                arr = json.load(f)
        else:
            arr = []
    except Exception:
        arr = []
    arr.append(submission)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(arr, f, ensure_ascii=False, indent=2)

    bot.send_message(message.chat.id, 
        '✅ Заявка принята!\n\n'
        '📞 Менеджер свяжется с вами:\n'
        f'{MANAGER_LINK}')
    user_states.pop(message.chat.id, None)


# ========== АДМИН КОМАНДЫ ==========

@bot.callback_query_handler(func=lambda call: call.data == 'admin_menu')
def admin_menu_handler(call):
    admin_id = call.message.chat.id
    
    if not is_admin(admin_id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    text = (
        '⚙️ <b>АДМИН ПАНЕЛЬ</b> ⚙️\n\n'
        '🔐 <b>Главное меню администратора</b>\n\n'
        '📋 Доступные функции:\n'
        '  • 💰 Управление балансом\n'
        '  • 👥 Управление пользователями\n'
        '  • 📋 Просмотр всех пользователей\n'
        '  • 📊 Статистика системы\n'
        '  • 🔍 Изучение пользователей\n\n'
        '⬇️ Выберите действие ниже:'
    )
    try:
        bot.edit_message_text(text, admin_id, call.message.message_id, reply_markup=admin_menu_kbd(admin_id), parse_mode='HTML')
    except:
        bot.send_message(admin_id, text, reply_markup=admin_menu_kbd(admin_id), parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_exit')
def admin_exit_handler(call):
    if not is_admin(call.message.chat.id):
        return
    send_menu(call.message.chat.id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_balance')
def admin_balance_handler(call):
    admin_id = call.message.chat.id
    
    if not is_admin(admin_id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    # Проверяем что это полный админ
    if is_support_admin(admin_id):
        bot.answer_callback_query(call.id, '❌ Эта функция недоступна для администраторов поддержки!', show_alert=True)
        return
    
    text = (
        '                                         \n'
        '   💰 УПРАВЛЕНИЕ БАЛАНСОМ 💰     \n'
        '                                         \n\n'
        'Выберите операцию:'
    )
    try:
        bot.edit_message_text(text, admin_id, call.message.message_id, reply_markup=admin_balance_kbd())
    except:
        bot.send_message(admin_id, text, reply_markup=admin_balance_kbd())
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_recharge_balance')
def admin_recharge_balance_handler(call):
    if call.message.chat.id != ADMIN_ID:
        return
    
    user_states[call.message.chat.id] = 'admin_waiting_user_id_recharge'
    bot.send_message(call.message.chat.id, '👤 Введите ID пользователя для пополнения:')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_add_balance')
def admin_add_balance_handler(call):
    if call.message.chat.id != ADMIN_ID:
        return
    
    user_states[call.message.chat.id] = 'admin_waiting_user_id_add'
    bot.send_message(call.message.chat.id, '👤 Введите ID пользователя:')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_remove_balance')
def admin_remove_balance_handler(call):
    if call.message.chat.id != ADMIN_ID:
        return
    
    user_states[call.message.chat.id] = 'admin_waiting_user_id_remove'
    bot.send_message(call.message.chat.id, '👤 Введите ID пользователя:')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_set_balance')
def admin_set_balance_handler(call):
    if call.message.chat.id != ADMIN_ID:
        return
    
    user_states[call.message.chat.id] = 'admin_waiting_user_id_set'
    bot.send_message(call.message.chat.id, '👤 Введите ID пользователя:')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_exploration')
def admin_exploration_handler(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    user_states[call.message.chat.id] = 'admin_checking_exploration'
    bot.send_message(call.message.chat.id, '👤 Введите ID пользователя для проверки статистики изучения:')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_all_users')
def admin_all_users_handler(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    users = load_users()
    
    if not users:
        text = (
            '                                         \n'
            '   📋 ВСЕ ПОЛЬЗОВАТЕЛИ 📋           \n'
            '                                         \n\n'
            '❌ <b>Нет зарегистрированных пользователей</b>'
        )
        try:
            bot.edit_message_text(text, call.message.chat.id, call.message.message_id, parse_mode='HTML')
        except:
            bot.send_message(call.message.chat.id, text, parse_mode='HTML')
        bot.answer_callback_query(call.id)
        return
    
    text = (
        '                                         \n'
        '   📋 ВСЕ ПОЛЬЗОВАТЕЛИ 📋           \n'
        '                                         \n\n'
        f'📊 <b>Всего пользователей:</b> <b>{len(users)}</b>\n\n'
        '<b> </b>\n\n'
    )
    
    for user_id, user_data in sorted(users.items()):
        first_name = user_data.get('first_name', 'Unknown')
        username = user_data.get('username', 'unknown')
        game_id = user_data.get('game_id', 'не указан')
        registered = user_data.get('registered', 'unknown')
        
        # Форматируем дату
        try:
            reg_date = registered.split('T')[0] if 'T' in registered else registered
        except:
            reg_date = registered
        
        text += (
            f'👤 <b>ID Telegram:</b> <code>{user_id}</code>\n'
            f'🎮 <b>ID в клубе:</b> <code>{game_id}</code>\n'
            f'👤 <b>Ник:</b> @{username}\n'
            f'📝 <b>Имя:</b> {first_name}\n'
            f'📅 <b>Регистрация:</b> {reg_date}\n'
            f'<b>━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━</b>\n\n'
        )
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=admin_menu_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=admin_menu_kbd(call.message.chat.id), parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_users')
def admin_users_handler(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    users = load_users()
    user_count = len(users)
    
    text = (
        '                                         \n'
        '   👥 УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ  \n'
        '                                         \n\n'
        f'📊 Всего пользователей: {user_count}\n\n'
        'Функции:\n'
        '  • Просмотр профилей\n'
        '  • Изменение баланса\n'
        '  • Управление рефералами\n'
    )
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=kb)
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=kb)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_stats')
def admin_stats_handler(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    users = load_users()
    referrals = load_referrals()
    balances = load_balances()
    
    total_users = len(users)
    total_referrals = sum(len(refs) for refs in referrals.values())
    total_balance = sum(balances.values())
    
    text = (
        '                                         \n'
        '   📊 СТАТИСТИКА СИСТЕМЫ 📊    \n'
        '                                         \n\n'
        f'👥 Всего пользователей: {total_users}\n'
        f'🔗 Всего рефералов: {total_referrals}\n'
        f'💰 Общий баланс: {total_balance}₽\n\n'
        f'⏰ Дата отчета: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    )
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=kb)
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=kb)
    bot.answer_callback_query(call.id)


# Обработчик текстовых сообщений для админ операций
@bot.message_handler(func=lambda message: user_states.get(message.chat.id, '') and 
                     (isinstance(user_states.get(message.chat.id), str) and user_states.get(message.chat.id, '').startswith('admin_') and user_states.get(message.chat.id, '') != 'admin_waiting_broadcast_message' or 
                      isinstance(user_states.get(message.chat.id), tuple)))
def admin_text_handler(message):
    if message.chat.id != ADMIN_ID:
        return
    
    state = user_states.get(message.chat.id)
    
    if state == 'admin_waiting_user_id_recharge':
        try:
            target_user_id = int(message.text)
            user_states[message.chat.id] = ('admin_waiting_amount_recharge', target_user_id)
            bot.send_message(message.chat.id, '💳 Введите сумму пополнения:')
        except:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите корректный ID')
    
    elif isinstance(state, tuple) and state[0] == 'admin_waiting_amount_recharge':
        try:
            target_user_id = state[1]
            amount = int(message.text)
            set_balance(target_user_id, amount)
            bot.send_message(message.chat.id, 
                f'✅ Пополнено успешно!\n\n'
                f'Пользователь: {target_user_id}\n'
                f'Сумма пополнения: +{amount}₽\n'
                f'Новый баланс: {amount}₽')
            user_states.pop(message.chat.id, None)
        except ValueError:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите сумму числом')
    
    elif state == 'admin_waiting_user_id_add':
        try:
            target_user_id = int(message.text)
            user_states[message.chat.id] = ('admin_waiting_amount_add', target_user_id)
            bot.send_message(message.chat.id, '💵 Введите сумму для добавления:')
        except:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите корректный ID')
    
    elif isinstance(state, tuple) and state[0] == 'admin_waiting_amount_add':
        try:
            target_user_id = state[1]
            amount = int(message.text)
            add_balance(target_user_id, amount)
            current_balance = get_balance(target_user_id)
            bot.send_message(message.chat.id, 
                f'✅ Успешно!\n\n'
                f'Пользователь: {target_user_id}\n'
                f'Добавлено: +{amount}₽\n'
                f'Новый баланс: {current_balance}₽')
            user_states.pop(message.chat.id, None)
        except ValueError:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите сумму числом')
    
    elif state == 'admin_waiting_user_id_remove':
        try:
            target_user_id = int(message.text)
            user_states[message.chat.id] = ('admin_waiting_amount_remove', target_user_id)
            bot.send_message(message.chat.id, '💵 Введите сумму для вычитания:')
        except:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите корректный ID')
    
    elif isinstance(state, tuple) and state[0] == 'admin_waiting_amount_remove':
        try:
            target_user_id = state[1]
            amount = int(message.text)
            add_balance(target_user_id, -amount)
            current_balance = get_balance(target_user_id)
            bot.send_message(message.chat.id, 
                f'✅ Успешно!\n\n'
                f'Пользователь: {target_user_id}\n'
                f'Вычтено: -{amount}₽\n'
                f'Новый баланс: {current_balance}₽')
            user_states.pop(message.chat.id, None)
        except ValueError:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите сумму числом')
    
    elif state == 'admin_waiting_user_id_set':
        try:
            target_user_id = int(message.text)
            user_states[message.chat.id] = ('admin_waiting_amount_set', target_user_id)
            bot.send_message(message.chat.id, '💵 Введите новый баланс:')
        except:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите корректный ID')
    
    elif isinstance(state, tuple) and state[0] == 'admin_waiting_amount_set':
        try:
            target_user_id = state[1]
            amount = int(message.text)
            set_balance(target_user_id, amount)
            bot.send_message(message.chat.id, 
                f'✅ Успешно!\n\n'
                f'Пользователь: {target_user_id}\n'
                f'Новый баланс: {amount}₽')
            user_states.pop(message.chat.id, None)
        except ValueError:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите сумму числом')
    
    elif state == 'admin_checking_exploration':
        try:
            target_user_id = int(message.text)
            stats = get_user_exploration_stats(target_user_id)
            
            # Создаем визуализацию процента
            filled = '█' * (stats['percent'] // 10)
            empty = '░' * (10 - (stats['percent'] // 10))
            progress_bar = f'[{filled}{empty}]'
            
            text = (
                '                                         \n'
                '   🔍 СТАТИСТИКА ИЗУЧЕНИЯ 🔍      \n'
                '                                         \n\n'
                f'👤 <b>Пользователь:</b> {target_user_id}\n'
                f'📊 <b>Изучил бота:</b> <b>{stats["percent"]}%</b>\n'
                f'{progress_bar}\n\n'
                f'🔘 <b>Нажато кнопок:</b> {stats["clicked"]} из {stats["total"]}\n\n'
                f'📈 <b>Подробно:</b>\n'
            )
            
            # Добавляем список нажатых кнопок
            if stats['actions']:
                for action, count in sorted(stats['actions'].items()):
                    button_emoji = {
                        'promos': '🎁',
                        'bonus': '🎖️',
                        'referral': '👥',
                        'copy_ref': '📋',
                        'ref_stats': '📊',
                        'profile': '👤',
                        'language': '🌐',
                        'menu': '🏠',
                        'lang_ru': '🇷🇺',
                        'lang_en': '🇬🇧',
                        'payment': '💳',
                        'community': '💬',
                        'manager': '👔'
                    }.get(action, '•')
                    text += f'{button_emoji} {action}: <b>{count}</b> нажатий\n'
            else:
                text += '(нет данных)'
            
            bot.send_message(message.chat.id, text, parse_mode='HTML')
            user_states.pop(message.chat.id, None)
        except ValueError:
            bot.send_message(message.chat.id, '❌ Ошибка! Введите корректный ID пользователя')


@bot.callback_query_handler(func=lambda call: call.data == 'admin_manage_admins')
def callbacks_admin_manage_admins(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admins = load_admins()
    admins_list = '\n'.join([f'🔑 ID: <code>{admin_id}</code>' for admin_id in admins])
    
    text = (
        '<b>🔑 УПРАВЛЕНИЕ АДМИНИСТРАТОРАМИ</b>\n\n'
        f'<b>Текущие администраторы ({len(admins)}):</b>\n'
        f'{admins_list}\n\n'
        '<b>Для добавления админа:</b> введите команду\n'
        '<code>/add_admin USER_ID</code>\n\n'
        '<b>Для удаления админа:</b> введите команду\n'
        '<code>/remove_admin USER_ID</code>'
    )
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, 
                            reply_markup=kb, parse_mode='HTML')
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=kb, parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_support_settings')
def callbacks_admin_support_settings(call):
    """Открывает настройки поддержки"""
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    enabled = is_support_enabled(admin_id)
    status = "✅ Включена" if enabled else "❌ Отключена"
    
    text = (
        f"📞 <b>УПРАВЛЕНИЕ ПОДДЕРЖКОЙ</b>\n\n"
        f"Статус: {status}\n\n"
        f"Когда поддержка <b>включена</b>, вы будете получать уведомления о запросах клиентов через /support\n\n"
        f"Тапните кнопку ниже, чтобы переключить статус:"
    )
    
    try:
        bot.edit_message_text(text, admin_id, call.message.message_id, 
                            reply_markup=support_settings_kbd(), parse_mode='HTML')
    except:
        bot.send_message(admin_id, text, reply_markup=support_settings_kbd(), parse_mode='HTML')
    
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_download_menu')
def callbacks_admin_download_menu(call):
    """Открывает меню скачивания"""
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    
    text = "📥 <b>СКАЧАТЬ ИНФОРМАЦИЮ</b>\n\nВыберите, что вы хотите скачать:"
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('👥 Информация о пользователях', callback_data='admin_download_users'))
    kb.add(types.InlineKeyboardButton('📊 Анализ пользователей', callback_data='admin_download_analysis'))
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    
    try:
        bot.edit_message_text(text, admin_id, call.message.message_id, 
                            reply_markup=kb, parse_mode='HTML')
    except:
        bot.send_message(admin_id, text, reply_markup=kb, parse_mode='HTML')
    
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_download_users')
def callbacks_admin_download_users(call):
    """Скачивает информацию о пользователях"""
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    
    bot.answer_callback_query(call.id, '⏳ Генерирую файл...', show_alert=False)
    
    try:
        excel_file = create_users_excel()
        
        bot.send_document(
            admin_id,
            excel_file,
            visible_file_name='users_info.xlsx',
            caption='📋 Информация о пользователях'
        )
        
        bot.send_message(admin_id, '✅ Файл загружен успешно!', reply_markup=admin_menu_kbd(admin_id))
    except Exception as e:
        bot.send_message(admin_id, f'❌ Ошибка при создании файла: {str(e)}', reply_markup=admin_menu_kbd(admin_id))


@bot.callback_query_handler(func=lambda call: call.data == 'admin_download_analysis')
def callbacks_admin_download_analysis(call):
    """Скачивает анализ пользователей"""
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    
    bot.answer_callback_query(call.id, '⏳ Генерирую файл...', show_alert=False)
    
    try:
        excel_file = create_analysis_excel()
        
        bot.send_document(
            admin_id,
            excel_file,
            visible_file_name='users_analysis.xlsx',
            caption='📊 Анализ пользователей'
        )
        
        bot.send_message(admin_id, '✅ Файл загружен успешно!', reply_markup=admin_menu_kbd(admin_id))
    except Exception as e:
        bot.send_message(admin_id, f'❌ Ошибка при создании файла: {str(e)}', reply_markup=admin_menu_kbd(admin_id))


@bot.message_handler(commands=['payment'])
def cmd_payment(message):
    """Команда для открытия информации об оплате"""
    track_action(message.chat.id, 'payment')
    user_id = message.chat.id
    
    users = load_users()
    user_str = str(user_id)
    user_info = users.get(user_str, {})
    
    # Если нет game_id - генерируем новый
    game_id = user_info.get('game_id')
    if not game_id:
        game_id = generate_game_id()
        # Сохраняем game_id
        if user_str in users:
            users[user_str]['game_id'] = game_id
            save_users(users)
    
    # Создаем ссылку с новым форматом: SERVICE_NAME_game_id
    payment_url = f'https://t.me/sapayobot?start={SERVICE_NAME}_{game_id}'
    
    referrer_id = user_info.get('referrer_id')
    referrer_info = users.get(str(referrer_id), {}) if referrer_id else {}
    referrer_name = referrer_info.get('first_name', 'Администратор')
    referrer_username = referrer_info.get('username', 'N/A')
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('💳 Перейти на оплату', url=payment_url))
    kb.add(types.InlineKeyboardButton('🎁 Пополнить кодом', callback_data='use_promo_code'))
    kb.add(types.InlineKeyboardButton('⬅️ Назад в меню', callback_data='menu'))
    
    text = (
        "💳 <b>ОПЛАТА</b>\n\n"
        f"🎮 Ваш ID: <code>{game_id}</code>\n\n"
    )
    
    if referrer_id and referrer_id in users:
        text += (
            f"<b>Реферер:</b> <code>@{referrer_username}</code>\n"
            f"({referrer_name})\n\n"
        )
    
    text += "Нажмите кнопку ниже для перехода на страницу оплаты или введите промокод"
    
    bot.send_message(user_id, text, reply_markup=kb, parse_mode='HTML')


@bot.message_handler(commands=['cancel'])
def cmd_cancel(message):
    """Команда для админа - завершить диалог поддержки"""
    admin_id = message.chat.id
    
    # Проверяем, находится ли админ в чате поддержки
    state = user_states.get(admin_id, '')
    
    if not state.startswith('support_chat_'):
        bot.send_message(admin_id, "❌ Вы не находитесь в диалоге поддержки")
        return
    
    ticket_id = state.replace('support_chat_', '')
    tickets = load_support_tickets()
    
    if ticket_id not in tickets:
        bot.send_message(admin_id, "❌ Тикет не найден")
        user_states.pop(admin_id, None)
        return
    
    ticket = tickets[ticket_id]
    client_id = ticket['client_id']
    client_name = ticket['client_name']
    
    # Уведомляем клиента
    bot.send_message(
        client_id,
        "⚠️ <b>Администратор завершил диалог</b>\n\n"
        "Если у вас еще есть вопросы, используйте команду /support",
        parse_mode='HTML'
    )
    
    # Подтверждаем админу
    bot.send_message(admin_id, f"✅ Диалог с <b>{client_name}</b> завершен", parse_mode='HTML')
    
    # Удаляем состояния
    user_states.pop(admin_id, None)
    user_states.pop(client_id, None)
    
    # Обновляем статус тикета
    ticket['status'] = 'closed'
    save_support_tickets(tickets)


@bot.message_handler(commands=['add_admin'])
def cmd_add_admin(message):
    if message.chat.id != ADMIN_ID:
        bot.send_message(message.chat.id, '❌ Только главный администратор может добавлять админов!')
        return
    
    try:
        args = message.text.split()
        if len(args) < 2:
            bot.send_message(message.chat.id, 
                '❌ Используйте: /add_admin USER_ID\n\n'
                'Пример: /add_admin 123456789')
            return
        
        new_admin_id = int(args[1])
        
        if add_admin(new_admin_id):
            text = (
                f'✅ <b>Администратор добавлен!</b>\n\n'
                f'ID: <code>{new_admin_id}</code>\n'
                f'Статус: Активный администратор'
            )
        else:
            text = (
                f'ℹ️ <b>Администратор уже существует</b>\n\n'
                f'ID: <code>{new_admin_id}</code>'
            )
        
        bot.send_message(message.chat.id, text, parse_mode='HTML')
    except ValueError:
        bot.send_message(message.chat.id, '❌ Ошибка! ID должен быть числом')


@bot.message_handler(commands=['remove_admin'])
def cmd_remove_admin(message):
    if message.chat.id != ADMIN_ID:
        bot.send_message(message.chat.id, '❌ Только главный администратор может удалять админов!')
        return
    
    try:
        args = message.text.split()
        if len(args) < 2:
            bot.send_message(message.chat.id, 
                '❌ Используйте: /remove_admin USER_ID\n\n'
                'Пример: /remove_admin 123456789')
            return
        
        admin_id = int(args[1])
        
        if admin_id == ADMIN_ID:
            bot.send_message(message.chat.id, '❌ Нельзя удалить главного администратора!')
            return
        
        if remove_admin(admin_id):
            text = (
                f'✅ <b>Администратор удален!</b>\n\n'
                f'ID: <code>{admin_id}</code>\n'
                f'Статус: Доступ отозван'
            )
        else:
            text = (
                f'ℹ️ <b>Администратор не найден</b>\n\n'
                f'ID: <code>{admin_id}</code>'
            )
        
        bot.send_message(message.chat.id, text, parse_mode='HTML')
    except ValueError:
        bot.send_message(message.chat.id, '❌ Ошибка! ID должен быть числом')


@bot.callback_query_handler(func=lambda call: call.data == 'admin_broadcast')
def callbacks_admin_broadcast(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    user_states[call.message.chat.id] = 'admin_waiting_broadcast_message'
    
    text = (
        '📢 <b>РЕЖИМ РАССЫЛКИ</b>\n\n'
        'Введите сообщение, которое будет отправлено всем пользователям:\n\n'
        '💡 <b>Советы:</b>\n'
        '  • Вы можете использовать HTML форматирование\n'
        '  • Сообщение будет отправлено в течение нескольких минут\n'
        '  • Не забудьте проверить текст перед отправкой!'
    )
    
    bot.send_message(call.message.chat.id, text, parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'admin_waiting_broadcast_message')
def process_broadcast_message(message):
    if not is_admin(message.chat.id):
        bot.send_message(message.chat.id, '❌ Доступ запрещен!')
        return
    
    admin_id = message.chat.id
    broadcast_text = message.text
    
    # Подтверждение перед отправкой
    preview_text = (
        '📋 <b>ПРЕДПРОСМОТР СООБЩЕНИЯ:</b>\n\n'
        f'{broadcast_text}\n\n'
        '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n'
        '✅ Подтвердите отправку рассылки всем пользователям?'
    )
    
    kb = types.InlineKeyboardMarkup()
    kb.add(
        types.InlineKeyboardButton('✅ Отправить', callback_data='admin_confirm_broadcast'),
        types.InlineKeyboardButton('❌ Отменить', callback_data='admin_cancel_broadcast')
    )
    
    # Сохраняем сообщение в user_states для последующей отправки
    user_states[admin_id] = ('admin_broadcast_ready', broadcast_text)
    
    bot.send_message(admin_id, preview_text, reply_markup=kb, parse_mode='HTML')


@bot.callback_query_handler(func=lambda call: call.data == 'admin_confirm_broadcast')
def callbacks_confirm_broadcast(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    state = user_states.get(admin_id)
    
    if not isinstance(state, tuple) or state[0] != 'admin_broadcast_ready':
        bot.answer_callback_query(call.id, '❌ Ошибка! Сообщение не найдено', show_alert=True)
        return
    
    broadcast_text = state[1]
    users = load_users()
    
    # Отправляем сообщение всем пользователям
    sent_count = 0
    failed_count = 0
    
    bot.send_message(admin_id, 
        '⏳ <b>Рассылка начата...</b>\n\n'
        'Пожалуйста, подождите, сообщение отправляется всем пользователям.',
        parse_mode='HTML')
    
    for user_id_str in users.keys():
        try:
            user_id = int(user_id_str)
            bot.send_message(user_id, broadcast_text, parse_mode='HTML')
            sent_count += 1
            time.sleep(0.1)  # Небольшая задержка чтобы не перегрузить Telegram API
        except Exception as e:
            failed_count += 1
    
    # Отправляем отчет
    report_text = (
        '✅ <b>РАССЫЛКА ЗАВЕРШЕНА!</b>\n\n'
        f'📤 Отправлено: <b>{sent_count}</b> сообщений\n'
        f'❌ Ошибок: <b>{failed_count}</b>\n'
        f'👥 Всего пользователей: <b>{len(users)}</b>\n\n'
        f'⏰ Время: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
    )
    
    bot.send_message(admin_id, report_text, reply_markup=admin_menu_kbd(admin_id), parse_mode='HTML')
    user_states.pop(admin_id, None)


@bot.callback_query_handler(func=lambda call: call.data == 'admin_cancel_broadcast')
def callbacks_cancel_broadcast(call):
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    user_states.pop(call.message.chat.id, None)
    
    text = '❌ <b>Рассылка отменена</b>'
    
    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, 
                            reply_markup=admin_menu_kbd(call.message.chat.id), parse_mode='HTML')
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=admin_menu_kbd(call.message.chat.id), parse_mode='HTML')
    
    bot.answer_callback_query(call.id, 'Рассылка отменена', show_alert=False)


# ========== СИСТЕМА ПОДДЕРЖКИ ==========

def create_users_excel():
    """Создает Excel файл с информацией о пользователях"""
    users = load_users()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Заголовки
    headers = ['ID пользователя', 'Имя', 'Username', 'Игровой ID', 'Баланс', 'Дата регистрации']
    ws.append(headers)
    
    # Стиль заголовка
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Добавляем данные
    for user_id, user_info in users.items():
        balance = get_balance(int(user_id))
        row = [
            user_id,
            user_info.get('first_name', 'N/A'),
            user_info.get('username', 'N/A'),
            user_info.get('game_id', '-'),
            balance,
            user_info.get('registered', 'N/A')
        ]
        ws.append(row)
    
    # Выравнивание столбцов
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    
    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def create_analysis_excel():
    """Создает Excel файл с анализом пользователей"""
    users = load_users()
    actions = load_user_actions()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анализ пользователей"
    
    # Заголовки
    headers = ['ID пользователя', 'Имя', 'Username', 'Баланс', 'Дата регистрации', 'Кликов за неделю', 'Всего кликов']
    ws.append(headers)
    
    # Стиль заголовка
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Рассчитываем дату неделю назад
    week_ago = datetime.now() - timedelta(days=7)
    
    # Добавляем данные
    for user_id, user_info in users.items():
        balance = get_balance(int(user_id))
        
        # Получаем действия пользователя
        user_actions = actions.get(str(user_id), [])
        
        # Считаем клики за неделю
        clicks_week = 0
        for action in user_actions:
            try:
                action_time = datetime.fromisoformat(action.get('timestamp', ''))
                if action_time > week_ago:
                    clicks_week += 1
            except:
                pass
        
        total_clicks = len(user_actions)
        
        row = [
            user_id,
            user_info.get('first_name', 'N/A'),
            user_info.get('username', 'N/A'),
            balance,
            user_info.get('registered', 'N/A'),
            clicks_week,
            total_clicks
        ]
        ws.append(row)
    
    # Выравнивание столбцов
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def support_settings_kbd():
    """Клавиатура для управления настройками поддержки админа"""
    kb = types.InlineKeyboardMarkup()
    admin_id = str(ADMIN_ID)
    enabled = is_support_enabled(int(admin_id))
    status = "✅ Включена" if enabled else "❌ Отключена"
    
    kb.add(types.InlineKeyboardButton(f'📞 Поддержка: {status}', callback_data='support_toggle'))
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    return kb


@bot.message_handler(commands=['support'])
def cmd_support(message):
    """Команда для клиентов связи со службой поддержки"""
    chat_id = message.chat.id
    current_time = time.time()
    
    # Проверка cooldown - защита от спама
    if chat_id in support_requests_cooldown:
        last_request = support_requests_cooldown[chat_id]
        time_passed = current_time - last_request
        
        if time_passed < SUPPORT_COOLDOWN_SECONDS:
            wait_time = int(SUPPORT_COOLDOWN_SECONDS - time_passed)
            bot.send_message(
                chat_id,
                f"⏳ Слишком частые запросы! Подождите {wait_time} секунд перед новым запросом."
            )
            return
    
    # Проверка активных тикетов
    if user_states.get(chat_id) == 'in_support':
        bot.send_message(chat_id, "⏳ Вы уже в диалоге с поддержкой. Дождитесь ответа.")
        return
    
    tickets = load_support_tickets()
    active_user_tickets = sum(1 for t in tickets.values() if t['client_id'] == chat_id and t['status'] in ['waiting', 'active'])
    
    if active_user_tickets >= MAX_ACTIVE_TICKETS_PER_USER:
        bot.send_message(chat_id, "❌ У вас уже есть активный запрос в поддержку. Завершите его перед новым запросом.")
        return
    
    # Обновляем cooldown
    support_requests_cooldown[chat_id] = current_time
    
    user_states[chat_id] = 'waiting_support_message'
    bot.send_message(chat_id, "📝 Опишите вашу проблему или вопрос:")


@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'waiting_support_message')
def process_support_request(message):
    """Обработка запроса поддержки от клиента"""
    chat_id = message.chat.id
    client_message = message.text
    
    # Защита от длинных сообщений (DDoS)
    if len(client_message) > 1000:
        bot.send_message(chat_id, "❌ Сообщение слишком длинное (максимум 1000 символов)")
        return
    
    # Загружаем админов
    admins = load_admins()
    
    # Фильтруем админов, у которых включена поддержка
    active_admins = [admin_id for admin_id in admins if is_support_enabled(admin_id)]
    
    if not active_admins:
        bot.send_message(chat_id, "❌ К сожалению, сейчас нет доступных администраторов. Попробуйте позже.")
        user_states.pop(chat_id, None)
        return
    
    # Создаем тикет поддержки
    ticket_id = str(int(time.time()))
    tickets = load_support_tickets()
    
    tickets[ticket_id] = {
        'client_id': chat_id,
        'client_name': message.from_user.first_name or 'Клиент',
        'message': client_message,
        'admin_id': None,
        'status': 'waiting',
        'created_at': datetime.now().isoformat()
    }
    save_support_tickets(tickets)
    
    # Отправляем уведомление всем активным админам
    notification_text = (
        f"📞 <b>НОВЫЙ ЗАПРОС В ПОДДЕРЖКУ</b>\n\n"
        f"👤 Клиент: <b>{message.from_user.first_name or 'Клиент'}</b>\n"
        f"💬 Сообщение: <i>{client_message}</i>\n\n"
        f"🆔 Тикет: <code>{ticket_id}</code>"
    )
    
    for admin_id in active_admins:
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton(f'✅ Принять', callback_data=f'support_accept_{ticket_id}'))
        
        try:
            bot.send_message(admin_id, notification_text, reply_markup=kb, parse_mode='HTML')
        except:
            pass
    
    user_states[chat_id] = 'in_support'
    bot.send_message(chat_id, "✅ Ваш запрос отправлен! Ожидайте ответа администратора...", parse_mode='HTML')


@bot.callback_query_handler(func=lambda call: call.data.startswith('support_accept_'))
def callbacks_support_accept(call):
    """Админ принимает тикет поддержки"""
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    ticket_id = call.data.replace('support_accept_', '')
    
    tickets = load_support_tickets()
    
    if ticket_id not in tickets:
        bot.answer_callback_query(call.id, '❌ Тикет не найден', show_alert=True)
        return
    
    ticket = tickets[ticket_id]
    
    # Если уже есть админ - не даем принять другому
    if ticket['admin_id'] is not None and ticket['admin_id'] != admin_id:
        bot.answer_callback_query(call.id, '❌ Этот тикет уже принят другим админом', show_alert=True)
        return
    
    # Принимаем тикет
    ticket['admin_id'] = admin_id
    ticket['status'] = 'active'
    save_support_tickets(tickets)
    
    # Уведомляем админа
    client_id = ticket['client_id']
    admin_name = call.message.chat.first_name or 'Администратор'
    
    bot.edit_message_text(
        f"✅ <b>ВЫ ПРИНЯЛИ ТИКЕТ #{ticket_id}</b>\n\n"
        f"👤 Клиент: <b>{ticket['client_name']}</b>\n"
        f"💬 Сообщение: <i>{ticket['message']}</i>",
        admin_id, call.message.message_id, parse_mode='HTML'
    )
    
    user_states[admin_id] = f'support_chat_{ticket_id}'
    
    # Уведомляем клиента
    bot.send_message(
        client_id,
        f"✅ <b>Ваш запрос принял администратор {admin_name}</b>\n\n"
        f"💬 Администратор сейчас будет вам помогать. Отправляйте свои вопросы.",
        parse_mode='HTML'
    )
    
    bot.answer_callback_query(call.id, '✅ Тикет принят!', show_alert=False)


@bot.message_handler(func=lambda message: str(user_states.get(message.chat.id, '')).startswith('support_chat_'))
def process_support_chat(message):
    """Обработка сообщений админа в чате поддержки"""
    admin_id = message.chat.id
    
    # Защита от длинных сообщений
    if len(message.text) > 1000:
        bot.send_message(admin_id, "❌ Сообщение слишком длинное (максимум 1000 символов)")
        return
    
    # Извлекаем ticket_id из состояния
    state = user_states.get(admin_id, '')
    if not state.startswith('support_chat_'):
        return
    
    ticket_id = state.replace('support_chat_', '')
    tickets = load_support_tickets()
    
    if ticket_id not in tickets:
        user_states.pop(admin_id, None)
        bot.send_message(admin_id, "❌ Тикет не найден. Чат закрыт.")
        return
    
    ticket = tickets[ticket_id]
    
    if ticket['admin_id'] != admin_id:
        bot.send_message(admin_id, "❌ Это не ваш чат с клиентом.")
        user_states.pop(admin_id, None)
        return
    
    client_id = ticket['client_id']
    
    # Отправляем сообщение админа клиенту с кнопкой подтверждения
    admin_name = message.from_user.first_name or 'Администратор'
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('✅ Хорошо, понятно', callback_data=f'support_ack_{ticket_id}'))
    
    try:
        bot.send_message(
            client_id,
            f"📨 <b>{admin_name}</b>:\n\n{message.text}",
            reply_markup=kb,
            parse_mode='HTML'
        )
    except:
        bot.send_message(admin_id, "❌ Ошибка отправки сообщения клиенту")
        return
    
    # Устанавливаем состояние клиента в ожидание действия
    user_states[client_id] = f'support_client_{ticket_id}'
    
    # Подтверждаем админу
    bot.send_message(admin_id, "✅ Сообщение отправлено клиенту")


@bot.message_handler(func=lambda message: str(user_states.get(message.chat.id, '')).startswith('support_client_') and not str(user_states.get(message.chat.id, '')).startswith('support_client_waiting_'))
def process_client_support_response(message):
    """Обработка ответа клиента в чате поддержки"""
    client_id = message.chat.id
    
    # Защита от длинных сообщений
    if len(message.text) > 1000:
        bot.send_message(client_id, "❌ Сообщение слишком длинное (максимум 1000 символов)")
        return
    
    # Извлекаем ticket_id из состояния
    state = user_states.get(client_id, '')
    if not state.startswith('support_client_'):
        return
    
    ticket_id = state.replace('support_client_', '')
    tickets = load_support_tickets()
    
    if ticket_id not in tickets:
        user_states.pop(client_id, None)
        bot.send_message(client_id, "❌ Диалог закрыт.")
        return
    
    ticket = tickets[ticket_id]
    admin_id = ticket['admin_id']
    client_name = ticket['client_name']
    
    # Отправляем сообщение клиента админу
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('💬 Продолжить чат', callback_data=f'support_continue_{ticket_id}'))
    kb.add(types.InlineKeyboardButton('✅ Закончить поддержку', callback_data=f'support_end_{ticket_id}'))
    
    try:
        bot.send_message(
            admin_id,
            f"💬 <b>{client_name}</b> написал:\n\n{message.text}",
            reply_markup=kb,
            parse_mode='HTML'
        )
    except:
        bot.send_message(client_id, "❌ Ошибка отправки. Попробуйте позже.")
        return
    
    # Подтверждаем клиенту
    bot.send_message(client_id, "✅ Ваше сообщение отправлено администратору")
    
    # Возвращаем клиента в режим ожидания
    user_states[client_id] = f'support_client_waiting_{ticket_id}'


@bot.callback_query_handler(func=lambda call: call.data.startswith('support_ack_'))
def callbacks_support_acknowledge(call):
    """Клиент нажимает кнопку 'Хорошо, понятно' или может отправить сообщение"""
    client_id = call.message.chat.id
    ticket_id = call.data.replace('support_ack_', '')
    
    tickets = load_support_tickets()
    
    if ticket_id not in tickets:
        bot.answer_callback_query(call.id, '❌ Диалог закрыт', show_alert=True)
        return
    
    ticket = tickets[ticket_id]
    admin_id = ticket['admin_id']
    client_name = ticket['client_name']
    
    # Устанавливаем состояние клиента в ожидание
    user_states[client_id] = f'support_client_waiting_{ticket_id}'
    
    # Отправляем уведомление админу
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('💬 Продолжить чат', callback_data=f'support_continue_{ticket_id}'))
    kb.add(types.InlineKeyboardButton('✅ Закончить поддержку', callback_data=f'support_end_{ticket_id}'))
    
    bot.send_message(
        admin_id,
        f"✅ Клиент <b>{client_name}</b> подтвердил получение сообщения",
        reply_markup=kb,
        parse_mode='HTML'
    )
    
    bot.edit_message_text(
        "✅ Сообщение получено\n\nВы можете написать ответ администратору или нажать кнопку ниже для завершения диалога",
        client_id, call.message.message_id
    )
    
    # Добавляем кнопку завершения для клиента
    kb_client = types.InlineKeyboardMarkup()
    kb_client.add(types.InlineKeyboardButton('✅ Завершить диалог', callback_data=f'support_client_end_{ticket_id}'))
    bot.send_message(client_id, "Выберите действие:", reply_markup=kb_client)
    
    bot.answer_callback_query(call.id, 'OK', show_alert=False)


@bot.callback_query_handler(func=lambda call: call.data.startswith('support_continue_'))
def callbacks_support_continue(call):
    """Админ продолжает чат"""
    admin_id = call.message.chat.id
    ticket_id = call.data.replace('support_continue_', '')
    
    user_states[admin_id] = f'support_chat_{ticket_id}'
    
    bot.edit_message_text(
        "💬 <b>Введите ваше сообщение для клиента:</b>",
        admin_id, call.message.message_id,
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id, 'Готово', show_alert=False)


@bot.callback_query_handler(func=lambda call: call.data.startswith('support_client_end_'))
def callbacks_support_client_end(call):
    """Клиент завершает диалог"""
    client_id = call.message.chat.id
    ticket_id = call.data.replace('support_client_end_', '')
    
    tickets = load_support_tickets()
    
    if ticket_id in tickets:
        ticket = tickets[ticket_id]
        admin_id = ticket['admin_id']
        client_name = ticket['client_name']
        
        # Уведомляем админа
        bot.send_message(
            admin_id,
            f"❌ <b>Клиент {client_name} завершил диалог</b>",
            parse_mode='HTML'
        )
    
    user_states.pop(client_id, None)
    
    bot.edit_message_text(
        "✅ <b>Диалог завершен</b>\n\n"
        "Спасибо за обращение! Если у вас еще есть вопросы, используйте команду /support",
        client_id, call.message.message_id,
        parse_mode='HTML'
    )
    
    bot.answer_callback_query(call.id, '✅ Диалог закрыт', show_alert=False)


@bot.callback_query_handler(func=lambda call: call.data.startswith('support_end_'))
def callbacks_support_end(call):
    """Админ завершает поддержку"""
    admin_id = call.message.chat.id
    ticket_id = call.data.replace('support_end_', '')
    
    tickets = load_support_tickets()
    
    if ticket_id in tickets:
        ticket = tickets[ticket_id]
        client_id = ticket['client_id']
        client_name = ticket['client_name']
        admin_name = call.message.chat.first_name or 'Администратор'
        
        # Уведомляем клиента
        bot.send_message(
            client_id,
            f"✅ <b>Диалог с {admin_name} завершен</b>\n\n"
            f"Спасибо за обращение! Если у вас еще есть вопросы, используйте команду /support",
            parse_mode='HTML'
        )
        
        # Удаляем тикет
        ticket['status'] = 'closed'
        save_support_tickets(tickets)
    
    user_states.pop(admin_id, None)
    
    bot.edit_message_text(
        "✅ <b>Поддержка завершена</b>",
        admin_id, call.message.message_id,
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id, '✅ Чат закрыт', show_alert=False)


@bot.callback_query_handler(func=lambda call: call.data == 'support_toggle')
def callbacks_support_toggle(call):
    """Админ включает/отключает поддержку"""
    if not is_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    admin_id = call.message.chat.id
    settings = load_support_settings()
    
    # Переключаем статус
    current_status = settings.get(str(admin_id), False)
    settings[str(admin_id)] = not current_status
    save_support_settings(settings)
    
    new_status = settings[str(admin_id)]
    status_text = "✅ Включена" if new_status else "❌ Отключена"
    
    bot.edit_message_text(
        f"📞 <b>Поддержка: {status_text}</b>\n\n"
        f"Вы {'начали' if new_status else 'прекратили'} принимать запросы клиентов",
        call.message.chat.id, call.message.message_id,
        reply_markup=support_settings_kbd(),
        parse_mode='HTML'
    )
    
    bot.answer_callback_query(call.id, f'Поддержка {status_text}', show_alert=False)


# Система приема чеков на зачисление денег
@bot.message_handler(commands=['receipt'])
def cmd_receipt(message):
    """Команда для приема чека в одну строку: /receipt username amount"""
    admin_id_setting = int(get_setting('ADMIN_ID', ADMIN_ID))
    receipt_agent_id = int(get_setting('RECEIPT_AGENT_ID', ADMIN_ID))
    
    if message.chat.id not in [admin_id_setting, receipt_agent_id]:
        bot.send_message(message.chat.id, "❌ Доступ запрещен! Команда только для администратора или агента.")
        return
    
    # Парсим аргументы: /receipt username amount
    args = message.text.split()
    
    if len(args) < 3:
        bot.send_message(
            message.chat.id,
            "📋 <b>НЕПРАВИЛЬНЫЙ ФОРМАТ КОМАНДЫ</b>\n\n"
            "Используйте: <code>/receipt username сумма</code>\n\n"
            "Пример: <code>/receipt GRBTMc 1000</code>",
            parse_mode='HTML'
        )
        return
    
    username = args[1].lstrip('@')
    
    try:
        amount = float(args[2])
        if amount <= 0:
            bot.send_message(message.chat.id, "❌ Сумма должна быть больше нуля!")
            return
    except ValueError:
        bot.send_message(message.chat.id, "❌ Пожалуйста, введите число (например: 1000 или 1000.50)")
        return
    
    # Ищем пользователя по username
    users = load_users()
    client_id = None
    
    for uid, user_info in users.items():
        if user_info.get('username', '').lower() == username.lower():
            client_id = int(uid)
            break
    
    if not client_id:
        bot.send_message(message.chat.id, f"❌ Пользователь с username @{username} не найден!")
        return
    
    # Находим реферера клиента
    client_info = users.get(str(client_id), {})
    client_game_id = client_info.get('game_id', 'N/A')
    referrer_id = client_info.get('referrer_id')
    
    # Сохраняем чек
    receipts = load_receipts()
    receipt_id = str(int(time.time()))
    
    # Применяем реферальную программу только если сумма больше 50000 рублей
    referral_threshold = int(get_setting('REFERRAL_THRESHOLD', 50000))
    if amount >= referral_threshold:
        referral_amount = amount * REFERRAL_PERCENT
    else:
        referral_amount = 0
    
    receipts[receipt_id] = {
        'client_id': client_id,
        'client_username': username,
        'client_game_id': client_game_id,
        'amount': amount,
        'referrer_id': referrer_id,
        'referral_amount': referral_amount,
        'created_at': datetime.now().isoformat(),
        'status': 'pending'
    }
    save_receipts(receipts)
    
    # Отправляем уведомление админу о необходимом переводе
    if referrer_id and referral_amount > 0:
        referrer_info = users.get(str(referrer_id), {})
        referrer_name = referrer_info.get('first_name', 'Клиент')
        referrer_id_display = referrer_id
        
        notification_text = (
            f"💰 <b>НОВЫЙ ЧЕК НА ЗАЧИСЛЕНИЕ</b>\n\n"
            f"🎮 Game ID: <code>{client_game_id}</code>\n"
            f"💵 Сумма: <code>{amount}</code>\n"
            f"📊 10% для реферера: <code>{referral_amount}</code>\n\n"
            f"<b>Реферер:</b> {referrer_name}\n"
            f"<b>ID для перевода:</b> <code>{referrer_id_display}</code>\n\n"
            f"🆔 ID чека: <code>{receipt_id}</code>"
        )
    else:
        notification_text = (
            f"💰 <b>НОВЫЙ ЧЕК НА ЗАЧИСЛЕНИЕ</b>\n\n"
            f"🎮 Game ID: <code>{client_game_id}</code>\n"
            f"💵 Сумма: <code>{amount}</code>\n"
            f"📊 Реферальная программа: "
        )
        if amount < referral_threshold:
            notification_text += f"Минимум {referral_threshold} RUB для активации\n"
        else:
            notification_text += f"Реферер не найден\n"
        notification_text += f"🆔 ID чека: <code>{receipt_id}</code>"
    
    admin_id_setting_int = int(get_setting('ADMIN_ID', ADMIN_ID))
    bot.send_message(admin_id_setting_int, notification_text, parse_mode='HTML')
    
    # Подтверждение админу, который ввел чек
    confirm_text = (
        f"✅ <b>ЧЕК ПРИНЯТ</b>\n\n"
        f"👤 Username: @{username}\n"
        f"🎮 Game ID: <code>{client_game_id}</code>\n"
        f"💵 Сумма: <code>{amount}</code>\n"
    )
    
    if amount >= referral_threshold and referrer_id:
        confirm_text += f"📊 10% реферу: <code>{referral_amount}</code>\n\n"
        confirm_text += f"Переведите <code>{referral_amount}</code> на ID <code>{referrer_id}</code>"
    else:
        if amount < referral_threshold:
            confirm_text += f"⚠️ Сумма меньше {referral_threshold} RUB - реферальная программа не активна"
        else:
            confirm_text += "⚠️ Реферер не найден"
    
    bot.send_message(message.chat.id, confirm_text, parse_mode='HTML')
    user_states.pop(message.chat.id, None)


# Система промокодов для пополнения баланса
@bot.callback_query_handler(func=lambda call: call.data == 'use_promo_code')
def callbacks_use_promo_code(call):
    """Открывает ввод промокода"""
    user_id = call.message.chat.id
    user_states[user_id] = 'waiting_promo_code'
    
    bot.send_message(
        user_id,
        "🎁 <b>ПОПОЛНЕНИЕ КОДОМ</b>\n\n"
        "Введите промокод:\n"
        "• <code>100 RUB</code>\n"
        "• <code>500 RUB</code>",
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'waiting_promo_code')
def process_promo_code(message):
    """Обработка промокода"""
    user_id = message.chat.id
    code = message.text.strip().upper()
    
    codes = load_promo_codes()
    
    if code not in codes:
        bot.send_message(user_id, "❌ Промокод не найден или неверный!", parse_mode='HTML')
        return
    
    code_info = codes[code]
    
    if code_info['status'] != 'active':
        bot.send_message(user_id, "❌ Этот промокод уже использован!", parse_mode='HTML')
        return
    
    # Применяем промокод
    amount = code_info['amount']
    
    # Добавляем баланс
    balances = load_balances()
    user_balance = float(balances.get(str(user_id), 0))
    user_balance += amount
    balances[str(user_id)] = user_balance
    save_balances(balances)
    
    # Помечаем код как использованный
    code_info['status'] = 'used'
    code_info['used_by'] = user_id
    code_info['used_at'] = datetime.now().isoformat()
    codes[code] = code_info
    save_promo_codes(codes)
    
    # Уведомляем админа
    users = load_users()
    user_info = users.get(str(user_id), {})
    username = user_info.get('username', 'unknown')
    
    admin_notification = (
        f"🎁 <b>ПРОМОКОД ИСПОЛЬЗОВАН</b>\n\n"
        f"👤 Пользователь: @{username}\n"
        f"💰 Сумма: <code>{amount} RUB</code>\n"
        f"🔑 Код: <code>{code}</code>\n"
        f"💵 Новый баланс: <code>{user_balance}</code>"
    )
    admin_id_setting = int(get_setting('ADMIN_ID', ADMIN_ID))
    bot.send_message(admin_id_setting, admin_notification, parse_mode='HTML')
    
    # Подтверждение пользователю
    confirmation = (
        f"✅ <b>ПРОМОКОД АКТИВИРОВАН</b>\n\n"
        f"💰 Баланс пополнен на: <code>{amount} RUB</code>\n"
        f"💵 Ваш новый баланс: <code>{user_balance} RUB</code>"
    )
    bot.send_message(user_id, confirmation, parse_mode='HTML')
    
    user_states.pop(user_id, None)
    track_action(user_id, f'use_promo_{amount}')


@bot.message_handler(commands=['generate_codes'])
def cmd_generate_codes(message):
    """Команда для админа - генерировать промокоды"""
    if message.chat.id != ADMIN_ID:
        bot.send_message(message.chat.id, "❌ Доступ запрещен! Команда только для администратора.")
        return
    
    user_states[message.chat.id] = 'waiting_code_amount'
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('100 RUB', callback_data='gen_code_100'))
    kb.add(types.InlineKeyboardButton('500 RUB', callback_data='gen_code_500'))
    
    bot.send_message(
        message.chat.id,
        "🎁 <b>ГЕНЕРАТОР ПРОМОКОДОВ</b>\n\n"
        "Выберите номинал кода:",
        reply_markup=kb,
        parse_mode='HTML'
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith('gen_code_'))
def callbacks_generate_codes(call):
    """Обработка выбора номинала"""
    amount = int(call.data.replace('gen_code_', ''))
    admin_id = call.message.chat.id
    
    user_states[admin_id] = {'gen_code_amount': amount, 'gen_code_step': 'waiting_quantity'}
    
    bot.send_message(
        admin_id,
        f"Сколько кодов на {amount} RUB нужно создать?\n\n"
        f"(Введите число, например: 10)"
    )
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda message: isinstance(user_states.get(message.chat.id), dict) and user_states.get(message.chat.id, {}).get('gen_code_step') == 'waiting_quantity')
def process_generate_codes_quantity(message):
    """Обработка количества кодов"""
    admin_id = message.chat.id
    
    try:
        quantity = int(message.text.strip())
        if quantity <= 0 or quantity > 100:
            bot.send_message(admin_id, "❌ Введите число от 1 до 100")
            return
    except ValueError:
        bot.send_message(admin_id, "❌ Введите корректное число")
        return
    
    state = user_states[admin_id]
    amount = state['gen_code_amount']
    
    # Создаем коды
    created_codes = create_promo_codes(amount, quantity)
    
    # Формируем текст с кодами
    codes_text = f"🎁 <b>СОЗДАНЫ КОДЫ НА {amount} RUB</b>\n\n"
    codes_text += f"Количество: <code>{quantity}</code>\n\n"
    codes_text += "<b>Коды:</b>\n"
    
    for code in created_codes:
        codes_text += f"<code>{code}</code>\n"
    
    bot.send_message(admin_id, codes_text, parse_mode='HTML')
    user_states.pop(admin_id, None)


# Система управления переменными конфигурации
@bot.callback_query_handler(func=lambda call: call.data == 'admin_settings_variables')
def callbacks_settings_variables(call):
    """Меню управления переменными"""
    if not is_full_admin(call.message.chat.id):
        bot.answer_callback_query(call.id, '❌ Доступ запрещен!', show_alert=True)
        return
    
    settings = load_settings()
    
    text = "⚙️ <b>НАСТРОЙКИ ПЕРЕМЕННЫХ</b>\n\n"
    text += f"<b>Текущие значения:</b>\n\n"
    text += f"🔑 <code>ADMIN_ID</code>: {settings.get('ADMIN_ID')}\n"
    text += f"📧 <code>RECEIPT_AGENT_ID</code>: {settings.get('RECEIPT_AGENT_ID')}\n"
    text += f"🏢 <code>SERVICE_NAME</code>: {settings.get('SERVICE_NAME')}\n"
    text += f"👥 <code>REFERRAL_PERCENT</code>: {settings.get('REFERRAL_PERCENT') * 100}%\n"
    text += f"💰 <code>REFERRAL_THRESHOLD</code>: {settings.get('REFERRAL_THRESHOLD')} RUB\n"
    text += f"⏱️ <code>SUPPORT_COOLDOWN</code>: {settings.get('SUPPORT_COOLDOWN_SECONDS')}сек\n"
    text += f"🎟️ <code>MAX_TICKETS</code>: {settings.get('MAX_ACTIVE_TICKETS_PER_USER')}\n"
    text += f"🖼️ <code>MENU_IMAGE</code>: {settings.get('MENU_IMAGE')}\n"
    text += f"🗑️ <code>AUTO_DELETE_TIMEOUT</code>: {settings.get('AUTO_DELETE_TIMEOUT')}сек\n\n"
    text += "Нажмите на параметр для изменения:"
    
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('🔑 ADMIN_ID', callback_data='set_var_ADMIN_ID'))
    kb.add(types.InlineKeyboardButton('📧 RECEIPT_AGENT_ID', callback_data='set_var_RECEIPT_AGENT_ID'))
    kb.add(types.InlineKeyboardButton('🏢 SERVICE_NAME', callback_data='set_var_SERVICE_NAME'))
    kb.add(types.InlineKeyboardButton('👥 REFERRAL_PERCENT', callback_data='set_var_REFERRAL_PERCENT'))
    kb.add(types.InlineKeyboardButton('💰 REFERRAL_THRESHOLD', callback_data='set_var_REFERRAL_THRESHOLD'))
    kb.add(types.InlineKeyboardButton('⏱️ SUPPORT_COOLDOWN', callback_data='set_var_SUPPORT_COOLDOWN_SECONDS'))
    kb.add(types.InlineKeyboardButton('🎟️ MAX_TICKETS', callback_data='set_var_MAX_ACTIVE_TICKETS_PER_USER'))
    kb.add(types.InlineKeyboardButton('🖼️ MENU_IMAGE', callback_data='set_var_MENU_IMAGE'))
    kb.add(types.InlineKeyboardButton('🗑️ AUTO_DELETE_TIMEOUT', callback_data='set_var_AUTO_DELETE_TIMEOUT'))
    kb.add(types.InlineKeyboardButton('⬅️ Назад', callback_data='admin_menu'))
    
    bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=kb, parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('set_var_'))
def callbacks_set_variable(call):
    """Запрашивает новое значение для переменной"""
    admin_id = call.message.chat.id
    var_name = call.data.replace('set_var_', '')
    
    user_states[admin_id] = {'edit_var': var_name, 'step': 'waiting_value'}
    
    current_value = get_setting(var_name)
    
    text = f"⚙️ <b>ИЗМЕНИТЬ {var_name}</b>\n\n"
    text += f"Текущее значение: <code>{current_value}</code>\n\n"
    text += "Введите новое значение:"
    
    bot.send_message(admin_id, text, parse_mode='HTML')
    bot.answer_callback_query(call.id)


@bot.message_handler(func=lambda message: isinstance(user_states.get(message.chat.id), dict) and user_states.get(message.chat.id, {}).get('step') == 'waiting_value')
def process_set_variable(message):
    """Обработка нового значения переменной"""
    admin_id = message.chat.id
    
    if not is_full_admin(admin_id):
        bot.send_message(admin_id, "❌ Доступ запрещен!")
        return
    
    state = user_states[admin_id]
    var_name = state['edit_var']
    new_value = message.text.strip()
    
    try:
        # Пытаемся конвертировать в нужный тип
        current_settings = load_settings()
        old_value = current_settings.get(var_name)
        
        # Определяем тип значения по текущему значению
        if isinstance(old_value, (int, float)):
            if var_name in ['REFERRAL_PERCENT']:
                new_value = float(new_value) / 100 if float(new_value) > 1 else float(new_value)
            else:
                new_value = int(new_value) if '.' not in new_value else float(new_value)
        
        set_setting(var_name, new_value)
        
        # Подтверждение
        confirmation = (
            f"✅ <b>ПАРАМЕТР ИЗМЕНЕН</b>\n\n"
            f"📝 {var_name}\n"
            f"❌ Было: <code>{old_value}</code>\n"
            f"✅ Стало: <code>{new_value}</code>"
        )
        
        bot.send_message(admin_id, confirmation, parse_mode='HTML')
        user_states.pop(admin_id, None)
        
    except ValueError:
        bot.send_message(admin_id, "❌ Неправильный формат значения! Попробуйте снова.")


if __name__ == '__main__':
    print('Bot is starting...')
    
    # Устанавливаем список команд
    commands = [
        telebot.types.BotCommand(command='start', description='Запустить бота'),
        telebot.types.BotCommand(command='support', description='Связаться со службой поддержки'),
        telebot.types.BotCommand(command='payment', description='Перейти к оплате'),
    ]
    bot.set_my_commands(commands)
    
    print('Commands registered')
    bot.infinity_polling()

